import streamlit as st
import pandas as pd
import fitz
import re
import numpy as np
import io
import json
import os
from datetime import datetime
from openpyxl.styles import numbers, Border, Side
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import firebase_admin
from firebase_admin import credentials, firestore
from concurrent.futures import ThreadPoolExecutor
import pdfplumber

# Convert Streamlit secret section to a dict manually
firebase_dict = {
    "type": st.secrets.firebase.type,
    "project_id": st.secrets.firebase.project_id,
    "private_key_id": st.secrets.firebase.private_key_id,
    "private_key": st.secrets.firebase.private_key,
    "client_email": st.secrets.firebase.client_email,
    "client_id": st.secrets.firebase.client_id,
    "auth_uri": st.secrets.firebase.auth_uri,
    "token_uri": st.secrets.firebase.token_uri,
    "auth_provider_x509_cert_url": st.secrets.firebase.auth_provider_x509_cert_url,
    "client_x509_cert_url": st.secrets.firebase.client_x509_cert_url,
}

# === Initialize Firebase ===
cred = credentials.Certificate(firebase_dict)

if not firebase_admin._apps:
    firebase_admin.initialize_app(cred)
db = firestore.client()

# === COMMON UTILITIES ===
standard_columns = ['PORT', '20', '40STD', '40HC', 'REMARKS']

@st.cache_data
def clean_numeric_series(col):
    return pd.to_numeric(
        col.astype(str)
        .str.replace(',', '')
        .str.replace('EUR', '')
        .str.replace('USD', '')
        .str.strip(),
        errors='coerce'
    )

def clean_numeric(series):
    def convert_to_numeric(value):
        if pd.isna(value) or value == '':
            return np.nan
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            # Remove currency symbols, commas, and extra whitespace
            cleaned = re.sub(r'[^\d.]', '', value.strip())
            try:
                return float(cleaned) if cleaned else np.nan
            except ValueError:
                return np.nan
        return np.nan

    return series.apply(convert_to_numeric)

def sanitize_text(text):
    return re.sub(r'[^\x20-\x7E]+', '', text)

def normalize_text(text):
    if pd.isna(text):
        return ''
    return ' '.join(str(text).lower().strip().split())

def standardize_pol(pol):
    """Standardize POL names to a consistent format."""
    if pd.isna(pol):
        return ''
    pol = str(pol).strip().lower()
    if 'nhava' in pol and 'sheva' in pol:
        return 'Nhava Sheva'
    return pol.title()

# === PARSING FUNCTIONS ===
def parse_turkon(file, month_year):
    try:
        records = []
        with pdfplumber.open(file) as pdf:
            for page in pdf.pages:
                table = page.extract_table({
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                    "snap_tolerance": 3,
                })

                if not table:
                    continue

                for row in table:
                    if not row or len(row) < 8:
                        continue
                    if 'POD' in row and '20' in row[4]:
                        continue  # skip header

                    pol = row[0]
                    sector = row[1]
                    country = row[2]
                    pod = row[3]
                    rate_20 = row[4]
                    rate_40hc = row[5]
                    rate_40rf = row[6] if row[6] != '-' else None
                    imo = row[7]

                    if not pol or not pod or not rate_20 or not imo:
                        continue

                    # Standardize POL: if POL is 'Nsa', display as 'Nhava Sheva'
                    pol_standardized = 'Nhava Sheva' if pol.strip().lower() == 'nsa' else pol.strip()

                    records.append({
                        'POL': pol_standardized,
                        'PORT': pod.strip(),
                        "20": rate_20.strip(),
                        "40HC": rate_40hc.strip(),
                        "40'HRF": rate_40rf.strip() if rate_40rf else None,
                        "IMO SC PER TEU": imo.strip()
                    })

        df = pd.DataFrame(records)
        df['POL'] = df['POL'].apply(standardize_pol)  # Standardize POL names

        # Split ports in the PORT column
        df = split_ports(df, port_column='PORT')

        return df, []
        
    except Exception as e:
        print(f"Error parsing Turkon file: {str(e)}")
        return pd.DataFrame(), [f"Error parsing Turkon file: {str(e)}"]

def parse_oocl(file, month_year):
    oocl_map = {'Origins*': 'POL', 'Destinations*': 'PORT', 'Rate 20': '20', 'Rate 40': '40STD', 'Rate 40H': '40HC'}
    valid_pols = {'Nhava Sheva', 'Mundra', 'Rajula'}
    sheets = ['Nhava Sheva', 'Mundra', 'Rajula']
    data = []
    for sheet in sheets:
        try:
            df = pd.read_excel(file, sheet_name=sheet)[list(oocl_map.keys())].rename(columns=oocl_map)
            df['POL'] = df['POL'].astype(str).str.strip().replace('nan', '')
            df['POL'] = df['POL'].apply(lambda x: next((pol for pol in valid_pols if pol.lower() in x.lower()), ''))
            df = df[df['POL'] != '']
            df = df[df['PORT'].notna()]
            # Clean numeric columns, remove decimal points, and add $ symbol
            df['20'] = clean_numeric(df['20']).apply(
                lambda x: f'${int(x)}' if pd.notna(x) and x != 0 else np.nan
            )
            df['40STD'] = clean_numeric(df['40STD']).apply(
                lambda x: f'${int(x)}' if pd.notna(x) and x != 0 else np.nan
            )
            df['40HC'] = clean_numeric(df['40HC']).apply(
                lambda x: f'${int(x)}' if pd.notna(x) and x != 0 else np.nan
            )
            data.append(df)
        except ValueError:
            continue

    # Concatenate all DataFrames from different sheets
    result_df = pd.concat(data, ignore_index=True) if data else pd.DataFrame()

    # Split ports in the PORT column
    result_df = split_ports(result_df, port_column='PORT')

    return result_df, []

def parse_emirates(file, month_year):
    try:
        # Load the Excel file to get sheet names
        xl = pd.ExcelFile(file)
        sheets = xl.sheet_names
        data = []
        all_terms = []

        def parse_sheet(sheet, file):
            try:
                df = pd.read_excel(file, sheet_name=sheet, header=None)
                header_row = 0
                for idx, row in df.iterrows():
                    if any(isinstance(cell, str) and 'Destination Port' in cell for cell in row):
                        header_row = idx
                        break
                if header_row == 0 and not any('Destination Port' in str(cell) for cell in df.values.flatten()):
                    return pd.DataFrame(), []
                terms_conditions = []
                terms_active = False
                # Add sheet name once at the beginning of the terms
                for idx, row in df.iterrows():
                    for cell in row:
                        if isinstance(cell, str) and 'Terms & Conditions' in cell:
                            terms_active = True
                            terms_conditions.append(f"Sheet: {sheet}")
                            continue
                        if terms_active and isinstance(cell, str) and cell.strip():
                            cell_text = cell.strip()
                            if re.match(r'^\d+\)', cell_text) or cell_text:
                                terms_conditions.append(cell_text)
                        elif terms_active and not any(isinstance(c, str) and c.strip() for c in row):
                            terms_active = False
                            break
                df = pd.read_excel(file, sheet_name=sheet, skiprows=header_row)
                df.columns = df.columns.str.strip()
                expected_cols = {
                    'Origin Port': 'POL',
                    'Destination Port': 'PORT',
                    'Service Name': 'SERVICE NAME',
                    'Routing': 'ROUTING',
                    'Transit\n time': 'TRANSIT TIME',
                    "20'GP": '20',
                    "40'HC": '40HC',
                    'Remarks': 'REMARKS',
                    'Notes Surcharges / Subject To': 'REMARKS'
                }
                selected_cols = {}
                for excel_col, new_name in expected_cols.items():
                    if isinstance(new_name, list):
                        for possible_name in new_name:
                            if possible_name in df.columns:
                                selected_cols[possible_name] = new_name[0]
                                break
                    elif excel_col in df.columns:
                        selected_cols[excel_col] = new_name
                if not all(key in selected_cols for key in ['Origin Port', 'Destination Port', "20'GP", "40'HC"]):
                    return pd.DataFrame(), terms_conditions
                df = df[list(selected_cols.keys())].rename(columns=selected_cols)
                header_row = df.columns[0]
                df = df[df['PORT'] != header_row]
                # Clean numeric columns and remove decimal points
                df['20'] = clean_numeric(df['20']).apply(
                    lambda x: str(int(x)) if pd.notna(x) and x != 0 else np.nan
                )
                df['40HC'] = clean_numeric(df['40HC']).apply(
                    lambda x: str(int(x)) if pd.notna(x) and x != 0 else np.nan
                )
                if 'REMARKS' in df.columns and 'Notes Surcharges / Subject To' in df.columns:
                    df['REMARKS'] = df['REMARKS'].fillna('') + ' ' + df['Notes Surcharges / Subject To'].fillna('')
                    df = df.drop(columns=['Notes Surcharges / Subject To'])
                elif 'Notes Surcharges / Subject To' in df.columns:
                    df = df.rename(columns={'Notes Surcharges / Subject To': 'REMARKS'})
                text_cols = ['SERVICE NAME', 'ROUTING', 'TRANSIT TIME', 'REMARKS', 'POL']
                for col in text_cols:
                    if col in df.columns:
                        df[col] = df[col].astype(str).str.strip().replace('nan', '') if col != 'TRANSIT TIME' else df[col].astype(str).str.strip().replace('-', np.nan).replace('nan', '')
                return df[df['PORT'].notna()], terms_conditions
            except Exception as e:
                print(f"Error parsing sheet {sheet}: {str(e)}")
                return pd.DataFrame(), [f"Error parsing sheet {sheet}: {str(e)}"]

        # Iterate over sheets
        for sheet in sheets:
            df, terms_conditions = parse_sheet(sheet, file)
            if not df.empty:
                data.append(df)
            all_terms.extend(terms_conditions)

        # Combine results
        result_df = pd.concat(data, ignore_index=True) if data else pd.DataFrame()

        # Split ports in the PORT column
        result_df = split_ports(result_df, port_column='PORT')

        return result_df, all_terms

    except Exception as e:
        print(f"Error parsing Emirates file: {str(e)}")
        return pd.DataFrame(), [f"Error parsing Emirates file: {str(e)}"]

def parse_hmm(file, month_year):
    try:
        # Read the main rate sheet
        df = pd.read_excel(file, sheet_name="HMM Rate Sheet ")
        df.columns = df.iloc[0]
        df = df.drop(index=0).reset_index(drop=True)
        df.columns = df.columns.str.strip()

        if 'POL CODE' not in df.columns or 'PORTS' not in df.columns:
            print("Warning: Required columns 'POL CODE' or 'PORTS' not found in HMM Rate Sheet.")
            return pd.DataFrame(), []

        forty_std_col = None
        forty_hc_col = None
        if '40STD' in df.columns:
            forty_std_col = '40STD'
        if '40DV/HC' in df.columns:
            forty_hc_col = '40DV/HC'

        if not (forty_std_col or forty_hc_col):
            print("Warning: Neither '40STD' nor '40DV/HC' found in HMM Rate Sheet.")
            return pd.DataFrame(), []

        cols_to_fetch = ['POL CODE', 'PORTS', '20DV']
        if forty_std_col:
            cols_to_fetch.append(forty_std_col)
        if forty_hc_col:
            cols_to_fetch.append(forty_hc_col)

        parsed = df[cols_to_fetch].copy()
        rename_map = {'POL CODE': 'POL', 'PORTS': 'PORT', '20DV': '20'}
        if forty_std_col:
            rename_map[forty_std_col] = '40STD'
        if forty_hc_col:
            rename_map[forty_hc_col] = '40HC'

        # Split ports in the PORTS column before renaming
        parsed = split_ports(parsed, port_column='PORTS')

        parsed = parsed.rename(columns=rename_map)
        parsed = parsed.dropna(subset=['POL', 'PORT'], how='all')  # Drop rows where both POL and PORT are NaN
        parsed['POL'] = parsed['POL'].astype(str).str.strip().replace('nan', '')
        parsed['POL'] = parsed['POL'].apply(standardize_pol)  # Standardize POL names
        parsed['PORT'] = parsed['PORT'].astype(str).str.strip().replace('nan', '')

        # Clean numeric columns and remove decimal points
        parsed['20'] = clean_numeric(parsed['20']).apply(
            lambda x: str(int(x)) if pd.notna(x) and x != 0 else np.nan
        )

        if '40STD' in parsed.columns:
            parsed['40STD'] = clean_numeric(parsed['40STD']).apply(
                lambda x: str(int(x)) if pd.notna(x) and x != 0 else np.nan
            )
        else:
            parsed['40STD'] = np.nan

        if '40HC' in parsed.columns:
            parsed['40HC'] = clean_numeric(parsed['40HC']).apply(
                lambda x: str(int(x)) if pd.notna(x) and x != 0 else np.nan
            )
        else:
            parsed['40HC'] = np.nan

        parsed = parsed[['POL', 'PORT', '20', '40STD', '40HC']]

        # Apply currency symbol without decimal points
        def apply_currency(value):
            if pd.notna(value) and value != 'nan':
                return f"{value} $"
            return value

        parsed['20'] = parsed['20'].apply(apply_currency)
        if '40STD' in parsed.columns:
            parsed['40STD'] = parsed['40STD'].apply(apply_currency)
        if '40HC' in parsed.columns:
            parsed['40HC'] = parsed['40HC'].apply(apply_currency)

        # Remove rows where all rate columns are NaN or "None"
        rate_columns = [col for col in ['20', '40STD', '40HC'] if col in parsed.columns]
        if rate_columns:
            parsed = parsed.dropna(subset=rate_columns, how='all')
            parsed = parsed[~parsed[rate_columns].eq('None').all(axis=1)]

        # Remove rows where PORT is empty or "None"
        parsed = parsed[parsed['PORT'].ne('') & parsed['PORT'].ne('None')]

        if '40STD' in parsed.columns and parsed['40STD'].isna().all():
            parsed = parsed.drop(columns=['40STD'])
        if '40HC' in parsed.columns and parsed['40HC'].isna().all():
            parsed = parsed.drop(columns=['40HC'])

        # Extract Terms and Conditions as Additional Information
        additional_info = []
        try:
            terms_df = pd.read_excel(file, sheet_name="Terms and Conditions", header=None)
            for idx, row in terms_df.iterrows():
                for cell in row:
                    if pd.notna(cell) and str(cell).strip():
                        additional_info.append(str(cell).strip())
        except Exception as e:
            print(f"Error reading Terms and Conditions sheet: {str(e)}")
            additional_info.append(f"Error reading Terms and Conditions sheet: {str(e)}")

        return parsed, additional_info

    except Exception as e:
        print(f"Error parsing HMM file: {str(e)}")
        return pd.DataFrame(), [f"Error parsing HMM file: {str(e)}"]
    
def parse_wan_hai(file, month_year):
    try:
        doc = fitz.open(stream=file.read(), filetype="pdf")
        text = "".join([page.get_text() for page in doc])

        # Updated regex to allow dash (-) and still capture the record
        matches = re.findall(r"([A-Z0-9 ,()'\-\w]{3,})\s+(\d{1,5})\s+(\d{1,5})\s+(\d{1,5})", text)

        # Create data list
        wan_hai_data = []
        for m in matches:
            port = m[0].strip()
            try:
                wan_hai_data.append((port, int(m[1]), int(m[2]), int(m[3])))
            except ValueError:
                # In case there is non-integer in numeric fields, skip
                pass

        # Create DataFrame
        df = pd.DataFrame(wan_hai_data, columns=['PORT', '20', '40STD', '40HC'])

        # Optionally add empty "REMARKS" column (was previously "Description")
        df['REMARKS'] = ''

        # Add POL column
        df.insert(0, 'POL', 'Nhava Sheva')
        df['POL'] = df['POL'].apply(standardize_pol)  # Standardize POL names

        # Split ports in the PORT column
        df = split_ports(df, port_column='PORT')

        return df, []
    except Exception as e:
        print(f"Error parsing Wan Hai file: {str(e)}")
        return pd.DataFrame(), []

def parse_one(file, month_year):
    data = []
    info = []
    try:
        # Load Excel file to get all sheet names
        xl = pd.ExcelFile(file)
        sheets = xl.sheet_names

        for sheet_name in sheets:
            sheet_name_lower = sheet_name.lower().strip()

            # Helper function to parse a table from a sheet
            def parse_table(df_raw, sheet_name, keywords, col_map, required_cols, is_latam=False):
                remarks = []
                data_frames = []
                current_row = 0

                while current_row < len(df_raw):
                    # Look for table start
                    data_start_row = None
                    for idx in range(current_row, len(df_raw)):
                        row = df_raw.iloc[idx]
                        row_str = ' '.join(row.dropna().astype(str).str.strip())
                        if not row_str.strip():  # Skip blank rows
                            continue
                        # Check for table start based on keywords
                        if sum(k in row_str.upper() for k in keywords) >= 2:
                            if idx + 1 < len(df_raw):
                                next_row = df_raw.iloc[idx + 1].dropna().astype(str).str.strip()
                                if not next_row.empty and (next_row.str.isnumeric().any() or next_row.str.contains(r'^[A-Za-z\s]+$', regex=True).any()):
                                    data_start_row = idx
                                    break
                        # Collect remarks without numbering or bullets
                        for cell in row:
                            if pd.notna(cell) and str(cell).strip():
                                remarks.append(str(cell).strip())
                    if data_start_row is None:
                        break

                    # Read table dynamically until an empty row or end of table
                    try:
                        # Determine the number of rows to read
                        max_rows = 150
                        for i in range(data_start_row + 1, min(len(df_raw), data_start_row + max_rows + 1)):
                            row = df_raw.iloc[i]
                            row_str = ' '.join(row.dropna().astype(str).str.strip())
                            if not row_str.strip():  # Stop at the first blank row
                                max_rows = i - data_start_row
                                break

                        df = pd.read_excel(file, sheet_name=sheet_name, skiprows=data_start_row, nrows=max_rows)
                        df.columns = df.columns.str.strip().str.replace("’", "'").str.replace("‘", "'")
                        print(f"{sheet_name} columns: {list(df.columns)}")

                        # Skip if the DataFrame is empty or has no valid columns
                        if df.empty or len(df.columns) == 0:
                            print(f"Skipping empty table in {sheet_name} at row {data_start_row}")
                            current_row = data_start_row + 1
                            continue

                        # For LATAM or LUX Service, handle remarks columns (combine REMARK and last column if both exist)
                        if is_latam and len(df.columns) > 1:
                            last_col = df.columns[-1]
                            remark_cols = [col for col in df.columns if col.lower() == 'remark']
                            if remark_cols and remark_cols[0] != last_col:
                                df['REMARKS'] = df[remark_cols[0]].combine_first(df[last_col])
                                df = df.drop(columns=[remark_cols[0], last_col])
                            else:
                                df = df.rename(columns={last_col: 'REMARKS'})
                            print(f"Renamed last column '{last_col}' to 'REMARKS' for {sheet_name}")

                        # Apply column mapping
                        mapped_cols = {}
                        used_targets = set()
                        for col in df.columns:
                            col_lower = col.lower().strip()
                            if is_latam and col == 'REMARKS':
                                mapped_cols[col] = 'REMARKS'
                                used_targets.add('REMARKS')
                                continue
                            for key, value in col_map.items():
                                if any(k in col_lower for k in key) and value not in used_targets:
                                    mapped_cols[col] = value
                                    used_targets.add(value)
                                    break

                        # Validate required columns
                        if not all(k in mapped_cols.values() for k in required_cols):
                            print(f"Missing required columns in {sheet_name}: {required_cols}, Mapped: {mapped_cols}")
                            current_row = data_start_row + max_rows
                            continue

                        parsed_df = df[list(mapped_cols.keys())].rename(columns=mapped_cols)

                        # Filter valid rows
                        if 'PORT' in parsed_df.columns:
                            parsed_df = parsed_df[~parsed_df['PORT'].astype(str).str.lower().isin(['port', 'ports', 'destination', 'dest'])]
                            parsed_df = parsed_df[~parsed_df['PORT'].astype(str).str.lower().str.contains('total|summary', na=False)]
                            parsed_df = parsed_df[parsed_df['PORT'].notna()]
                            parsed_df = parsed_df[parsed_df['PORT'].astype(str).str.strip() != '']
                            parsed_df['PORT'] = parsed_df['PORT'].astype(str).str.strip()
                        else:
                            print(f"PORT column missing in {sheet_name} after mapping")
                            current_row = data_start_row + max_rows
                            continue

                        # Clean numeric columns and remove decimal points
                        for col in ['20', '40STD', '40HC']:
                            if col in parsed_df.columns:
                                parsed_df[col] = clean_numeric(parsed_df[col]).apply(
                                    lambda x: str(int(x)) if pd.notna(x) and x != 0 else np.nan
                                )
                            else:
                                parsed_df[col] = np.nan

                        # Remove decimal points from TRANSIT TIME
                        if 'TRANSIT TIME' in parsed_df.columns:
                            parsed_df['TRANSIT TIME'] = parsed_df['TRANSIT TIME'].apply(
                                lambda x: str(int(float(x))) if pd.notna(x) and x != '' and float(x) != 0 else np.nan
                            )

                        # Exclude rows where all rate columns are NaN
                        rate_cols = [col for col in ['20', '40STD', '40HC'] if col in parsed_df.columns]
                        if rate_cols:
                            parsed_df = parsed_df[parsed_df[rate_cols].notna().any(axis=1)]
                        else:
                            print(f"No rate columns found in {sheet_name}, skipping NaN rate filter")

                        # Handle remarks: Apply the same remark to all rows in the table for LATAM or LUX Service
                        if is_latam and 'REMARKS' in parsed_df.columns:
                            if parsed_df['REMARKS'].notna().any():
                                remark_value = parsed_df['REMARKS'].dropna().iloc[0]
                            else:
                                remark_value = ''
                            parsed_df['REMARKS'] = remark_value
                            print(f"Applied remark '{remark_value}' to all rows in {sheet_name} table")
                        elif 'REMARKS' in parsed_df.columns:
                            parsed_df['REMARKS'] = parsed_df['REMARKS'].astype(str).str.strip().replace('nan', '')
                        else:
                            parsed_df['REMARKS'] = ''

                        # Add POL and Sheet
                        parsed_df.insert(0, 'POL', 'Nhava Sheva')
                        parsed_df['POL'] = parsed_df['POL'].apply(standardize_pol)
                        parsed_df['Sheet'] = sheet_name

                        # Ensure unique column names
                        parsed_df.columns = pd.Index([f"{col}_{i}" if parsed_df.columns.duplicated().any() else col 
                                                     for i, col in enumerate(parsed_df.columns)])

                        print(f"{sheet_name} parsed successfully: {parsed_df.shape}")
                        data_frames.append(parsed_df)

                        # Move current_row to the end of the parsed table
                        current_row = data_start_row + max_rows

                        # Add remarks to info with the new format
                        if remarks:
                            info.append(f"* Sheet : {sheet_name}")
                            info.extend(remarks)
                            info.append("")  # Add a blank line after each sheet's remarks
                    except Exception as e:
                        print(f"Error parsing table in {sheet_name} at row {data_start_row}: {str(e)}")
                        # If info doesn't already have the sheet name, add it
                        if not info or info[-1] != f"* Sheet : {sheet_name}":
                            info.append(f"* Sheet : {sheet_name}")
                        info.append(f"Error parsing table in {sheet_name} at row {data_start_row}: {str(e)}")
                        info.append("")  # Add a blank line after the error
                        current_row = data_start_row + 1

                return data_frames, remarks

            # === FAREAST and GULF ===
            if "fareast" in sheet_name_lower and "gulf" in sheet_name_lower:
                col_map = {
                    ('fpd', 'port', 'dest', 'destination'): 'PORT',
                    ("20'", '20d', '20dc', '20 dc', '20 ', "20'd"): '20',
                    ('40hc', 'hcd', '40 hc', '40high', '40 high', "40'hc"): '40HC',
                    ("40'", '40d', '40std', '40 std', '40 ', "40'd"): '40STD',
                    ('country',): 'POD COUNTRY',
                    ('remark', 'rate structure', 'surcharges group'): 'REMARKS'
                }
                keywords = ['FPD', 'PORT', 'DEST', 'DESTINATION', "20'", "40'", 'HCD']
                required_cols = ['PORT', '20', '40STD', '40HC']
                dfs, _ = parse_table(pd.read_excel(file, sheet_name=sheet_name, header=None), sheet_name, keywords, col_map, required_cols)
                data.extend(dfs)

            # === EUR and MED ===
            elif "eur" in sheet_name_lower and "med" in sheet_name_lower:
                col_map = {
                    ('del description',): 'PORT',
                    ('oft 20',): '20',
                    ('oft 40',): '40STD',
                    ('oft hc',): '40HC',
                    ('expiry date',): 'EXPIRY DATE',
                    ('remarks', 'rate structure', 'surcharges group', 'include surcharge'): 'REMARKS'
                }
                keywords = ['DEL DESCRIPTION', 'OFT 20', 'OFT 40', 'OFT HC']
                required_cols = ['PORT', '20', '40STD', '40HC']
                dfs, _ = parse_table(pd.read_excel(file, sheet_name=sheet_name, header=None), sheet_name, keywords, col_map, required_cols)
                for df in dfs:
                    if 'EXPIRY DATE' in df.columns:
                        df['EXPIRY DATE'] = pd.to_datetime(df['EXPIRY DATE'], errors='coerce').dt.strftime('%Y-%m-%d')
                data.extend(dfs)

            # === AUS and NZ ===
            elif "aus" in sheet_name_lower and "nz" in sheet_name_lower:
                col_map = {
                    ('port', 'dest', 'destination'): 'PORT',
                    ("20'", '20d', '20dc', '20 dc', '20 ', "20'd"): '20',
                    ('40hc', 'hcd', '40 hc', '40high', '40 high', "40'hc"): '40HC',
                    ("40'", '40d', '40std', '40 std', '40 ', "40'd"): '40STD',
                    ('remark', 'rate structure', 'surcharges group'): 'REMARKS',
                    ('t/t',): 'TRANSIT TIME'
                }
                keywords = ['PORT', "20'", "40'", "40'HC", 'DEST', 'DESTINATION']
                required_cols = ['PORT', '20', '40STD', '40HC']
                dfs, _ = parse_table(pd.read_excel(file, sheet_name=sheet_name, header=None), sheet_name, keywords, col_map, required_cols)
                data.extend(dfs)

            # === AFRICA ===
            elif "africa" in sheet_name_lower:
                col_map = {
                    ('port', 'dest', 'destination'): 'PORT',
                    ("20'", '20d', '20dc', '20 dc', '20 ', "20'd"): '20',
                    ('40hc', 'hcd', '40 hc', '40high', '40 high', "40'hc"): '40HC',
                    ("40'", '40d', '40std', '40 std', '40 ', "40'd"): '40STD',
                    ('remark', 'rate structure', 'surcharges group'): 'REMARKS',
                    ('t/t',): 'TRANSIT TIME'
                }
                keywords = ['PORT', "20'", "40'", "40'HC", 'REMARKS', 'DEST', 'DESTINATION']
                required_cols = ['PORT', '20', '40STD', '40HC']
                dfs, _ = parse_table(pd.read_excel(file, sheet_name=sheet_name, header=None), sheet_name, keywords, col_map, required_cols)
                data.extend(dfs)

            # === LATAM ===
            elif "latam" in sheet_name_lower:
                col_map = {
                    ('port', 'dest', 'destination', 'ports'): 'PORT',
                    ("20'", '20d', '20dc', '20 dc', '20 ', "20'd"): '20',
                    ('40hc', 'hcd', '40 hc', '40high', '40 high', "40'hc"): '40HC',
                    ("40'", '40d', '40std', '40 std', '40 ', "40'd"): '40STD',
                    ('remark', 'rate structure', 'surcharges group'): 'REMARKS',
                    ('t/t',): 'TRANSIT TIME'
                }
                keywords = ['PORT', "20'", "40'", "40'HC", 'REMARKS', 'DEST', 'DESTINATION', 'PORTS']
                required_cols = ['PORT', '20', '40STD', '40HC']
                dfs, _ = parse_table(pd.read_excel(file, sheet_name=sheet_name, header=None), sheet_name, keywords, col_map, required_cols, is_latam=True)
                data.extend(dfs)

            # === LUX Service for LAEC ===
            elif "lux" in sheet_name_lower and "laec" in sheet_name_lower:
                col_map = {
                    ('port', 'dest', 'destination', 'ports'): 'PORT',
                    ("20'", '20d', '20dc', '20 dc', '20 ', "20'd"): '20',
                    ('40hc', 'hcd', '40 hc', '40high', '40 high', "40'hc"): '40HC',
                    ("40'", '40d', '40std', '40 std', '40 ', "40'd"): '40STD',
                    ('remark', 'rate structure', 'surcharges group'): 'REMARKS'
                }
                keywords = ['PORT', "20'", "40'", "40'HC", 'REMARKS', 'DEST', 'DESTINATION', 'PORTS']
                required_cols = ['PORT', '20', '40STD', '40HC']
                dfs, _ = parse_table(pd.read_excel(file, sheet_name=sheet_name, header=None), sheet_name, keywords, col_map, required_cols, is_latam=True)
                data.extend(dfs)

        # Combine all data and remove duplicates
        if data:
            all_columns = set()
            for df in data:
                all_columns.update(df.columns)
            all_columns = sorted(list(all_columns))
            
            aligned_dfs = []
            for df in data:
                missing_cols = [col for col in all_columns if col not in df.columns]
                for col in missing_cols:
                    df[col] = np.nan
                aligned_dfs.append(df[all_columns])
            
            final_df = pd.concat(aligned_dfs, ignore_index=True)
            final_df = final_df.drop_duplicates(subset=['POL', 'PORT', '20', '40STD', '40HC', 'Sheet'], keep='first')

            # Define the desired column order with TRANSIT TIME and ROUTING
            desired_column_order = [
                'POL', 'PORT', '20', '40STD', '40HC', 'POD COUNTRY',
                'EXPIRY DATE', 'TRANSIT TIME', 'ROUTING', 'REMARKS', 'Sheet'
            ]

            # Reorder columns according to desired_column_order
            final_columns = [col for col in desired_column_order if col in final_df.columns]
            remaining_columns = [col for col in final_df.columns if col not in desired_column_order]
            final_columns.extend(remaining_columns)
            final_df = final_df[final_columns]
        else:
            final_df = pd.DataFrame()

        # Remove duplicates in info and filter short lines (but keep sheet names and blank lines)
        info = list(dict.fromkeys([line for line in info if len(line.split()) > 2 or line.startswith("* Sheet :") or line == ""]))
        return final_df, info

    except Exception as e:
        print(f"Error in parse_one: {str(e)}")
        # Ensure the error is added to info with a sheet context if possible
        if sheets and info and not info[-1].startswith("* Sheet :"):
            info.append(f"* Sheet : {sheets[-1]}")
        info.append(f"Error in parse_one: {str(e)}")
        return pd.DataFrame(), info



def parse_msc(file, month_year):
    try:
        import io
        import pdfplumber

        def extract_dataframe_from_pdf(pdf_file):
            all_rows = []
            with pdfplumber.open(pdf_file) as pdf:
                for page in pdf.pages:
                    table = page.extract_table()
                    if table:
                        all_rows.extend(table)
            return pd.DataFrame(all_rows)

        # Determine file type
        file_ext = str(file.name).lower()
        if file_ext.endswith('.pdf'):
            df = extract_dataframe_from_pdf(file)
        else:
            df = pd.read_excel(file, header=None)

        # Locate data start row
        data_start = 0
        for idx, row in df.iterrows():
            if isinstance(row.iloc[0], str) and 'SECTOR' in row.iloc[0].upper():
                data_start = idx + 1
                break

        if data_start == 0:
            return pd.DataFrame(), []

        df_data = df.iloc[data_start:].reset_index(drop=True)

        parsed = pd.DataFrame({
            'PORT': df_data.iloc[:, 0].dropna().astype(str).str.strip(),
            '20': df_data.iloc[:, 1],
            '40STD': df_data.iloc[:, 2],
            'REMARKS': df_data.iloc[:, 3].fillna('').astype(str).str.strip()
        })

        parsed = parsed[parsed['PORT'].str.strip() != ''].dropna(subset=['PORT'])
        desired_columns = ['PORT', '20', '40STD', 'REMARKS']
        parsed = parsed.dropna(subset=desired_columns)

        def extract_value_with_currency(value):
            if isinstance(value, str):
                value = value.strip()
                if 'EUR' in value.upper():
                    match = re.search(r'(\d+[.,]?\d*)', value)
                    if match:
                        num = float(match.group(1).replace(',', ''))
                        return f"{int(num)} €"
                else:
                    match = re.search(r'(\d+[.,]?\d*)', value)
                    if match:
                        num = float(match.group(1).replace(',', ''))
                        return f"{int(num)} $"
            elif isinstance(value, (int, float)):
                return f"{int(value)} $"
            return np.nan

        parsed['20'] = parsed['20'].apply(extract_value_with_currency)
        parsed['40STD'] = parsed['40STD'].apply(extract_value_with_currency)

        parsed.insert(0, 'POL', 'Nhava Sheva')
        parsed['POL'] = parsed['POL'].apply(standardize_pol)

        # Split ports in the PORT column
        parsed = split_ports(parsed, port_column='PORT')

        return parsed, []

    except Exception as e:
        print(f"Error in parse_msc: {str(e)}")
        return pd.DataFrame(), []

def parse_msc_eur_med(file, month_year):
    try:
        # Load the Excel file
        df = pd.read_excel(file, header=None)

        # Locate the header row by searching for "PORTS" in the first column
        header_row_idx = None
        for idx, row in df.iterrows():
            if isinstance(row.iloc[0], str) and 'PORTS' in row.iloc[0].upper():
                header_row_idx = idx
                break

        if header_row_idx is None:
            print("Could not find header row with 'PORTS' in the first column")
            return pd.DataFrame(), ["Header row not found"]

        # Extract remarks from rows above the header
        remarks = []
        for idx in range(header_row_idx):
            row = df.iloc[idx]
            for cell in row:
                if pd.notna(cell) and str(cell).strip():
                    remarks.append(str(cell).strip())

        # Since "20" and "40" are in the row below "RATE", use the row after "PORTS" as the header
        header_row = df.iloc[header_row_idx]
        sub_header_row = df.iloc[header_row_idx + 1]

        # Combine the header and sub-header to create proper column names
        combined_columns = []
        for main, sub in zip(header_row, sub_header_row):
            main = str(main).strip() if pd.notna(main) else ''
            sub = str(sub).strip() if pd.notna(sub) else ''
            if main.upper() == 'RATE' and sub in ['20', '40', '40.0']:
                combined_columns.append(sub)
            elif main.upper() in ['ETS', 'FUEL EU', 'VALIDITY', 'ANCILLARIES/ SURCHARGES']:
                combined_columns.append(main)
            elif main.upper() == 'PORTS':
                combined_columns.append(main)
            else:
                combined_columns.append(main if main else sub)

        # Set the data rows starting after the sub-header row
        df_data = df.iloc[header_row_idx + 2:].reset_index(drop=True)
        df_data.columns = combined_columns

        # Clean column names by stripping whitespace and standardizing
        df_data.columns = [str(col).strip().upper() for col in df_data.columns]
        

        # Rename columns to match expected format, handling variations like '40.0'
        column_mapping = {}
        for col in df_data.columns:
            col_upper = str(col).strip().upper()
            if col_upper == 'PORTS':
                column_mapping[col] = 'PORT'
            elif '20' in col_upper:
                column_mapping[col] = '20'
            elif '40' in col_upper or col_upper == '40.0':
                column_mapping[col] = '40STD'
            elif 'ANCILLARIES/ SURCHARGES' in col_upper:
                column_mapping[col] = 'REMARKS'
            elif 'ETS' in col_upper:
                column_mapping[col] = 'ETS'
            elif 'FUEL EU' in col_upper:
                column_mapping[col] = 'FUEL EU'
            elif 'VALIDITY' in col_upper:
                column_mapping[col] = 'VALIDITY'

        df_data = df_data.rename(columns=column_mapping)

        # Verify that required columns exist
        required_columns = ['PORT', '20', '40STD', 'REMARKS', 'ETS', 'FUEL EU', 'VALIDITY']
        missing_columns = [col for col in required_columns if col not in df_data.columns]
        if missing_columns:
            print(f"Missing columns after renaming: {missing_columns}")
            return pd.DataFrame(), remarks + [f"Missing columns: {missing_columns}"]

        # Filter out rows where PORT is empty or not a string
        parsed = df_data[df_data['PORT'].apply(lambda x: isinstance(x, str) and x.strip() != '')].copy()

        if parsed.empty:
            print("No valid port data found after filtering")
            return pd.DataFrame(), remarks + ["No valid port data found"]

        # Clean numeric columns (20 and 40STD), ensuring no decimal points
        def clean_numeric_series(col):
            def clean_value(val):
                if pd.isna(val) or val is None or str(val).strip() == '':
                    return np.nan
                if isinstance(val, (int, float)):
                    return int(val)  # Convert to integer to remove decimal points
                val_str = str(val).strip()
                val_clean = val_str.replace(',', '').replace('$', '').replace('USD', '').strip()
                if not val_clean.replace('.', '').replace('-', '').isdigit():
                    print(f"Invalid value in clean_numeric_series: '{val_str}'")
                    return np.nan
                try:
                    return int(float(val_clean))  # Convert to float first, then to int to remove decimal points
                except (ValueError, TypeError) as e:
                    print(f"Error converting value '{val_str}' to int: {str(e)}")
                    return np.nan
            return col.apply(clean_value)

        parsed['20'] = clean_numeric_series(parsed['20'])
        parsed['40STD'] = clean_numeric_series(parsed['40STD'])

        # Explicitly convert to Int64 type to ensure integer storage and display
        parsed['20'] = parsed['20'].astype('Int64')
        parsed['40STD'] = parsed['40STD'].astype('Int64')

        # Add default POL as Nhava Sheva
        parsed.insert(0, 'POL', 'Nhava Sheva')

        # Define standardize_pol function
        def standardize_pol(pol):
            if pd.isna(pol):
                return ''
            pol = str(pol).strip().upper()
            mapping = {
                'INNHV': 'Nhava Sheva',
                'NHAVA SHEVA': 'Nhava Sheva',
                'NHV/RQ': 'Nhava Sheva',
                'NHAVA': 'Nhava Sheva',
                'NSA': 'Nhava Sheva',
                'NS': 'Nhava Sheva',
                'MUNDRA': 'Mundra',
                'PIPAVAV': 'Pipavav',
                'TUTICORIN': 'Tut',
                'TUT': 'Tut',
                'KATTUPALLI': 'Ktp',
                'KOLKATA': 'Ccu',
                'CCU': 'Ccu',
                'WRG': 'Wrg',
                'INMUN': 'Mundra',
                'INNSA': 'Nhava Sheva',
                'HAZIRA': 'Hazira',
                'HZ': 'Hazira',
                'INHZA': 'Hazira'
            }
            return mapping.get(pol, pol.title())

        # Apply POL standardization
        parsed['POL'] = parsed['POL'].apply(standardize_pol)

        # Clean and standardize PORT names
        parsed['PORT'] = parsed['PORT'].astype(str).str.strip().str.title()

        # Define split_ports function to handle ports with slashes
        def split_ports(df, port_column='PORT'):
            def split_and_clean_port(port):
                if pd.isna(port) or not port:
                    return ['']
                port = str(port).strip()
                if '/' in port:
                    ports = [p.strip() for p in port.split('/')]
                else:
                    ports = [port]
                cleaned_ports = []
                for p in ports:
                    if p:
                        cleaned = p[0].upper() + p[1:].lower() if len(p) > 1 else p.upper()
                        cleaned_ports.append(cleaned)
                return cleaned_ports if cleaned_ports else ['']
            
            new_rows = []
            for idx, row in df.iterrows():
                ports = split_and_clean_port(row[port_column])
                for port in ports:
                    new_row = row.copy()
                    new_row[port_column] = port
                    new_rows.append(new_row)
            return pd.DataFrame(new_rows).reset_index(drop=True)

        # Split ports in the PORT column (e.g., "Sines&(ICD-BOBADELA)")
        parsed['PORT'] = parsed['PORT'].str.replace(r'&(ICD-BOBADELA)', '', regex=True)  # Remove extra info
        parsed = split_ports(parsed, port_column='PORT')

        # Clean REMARKS, ETS, FUEL EU, and VALIDITY
        parsed['REMARKS'] = parsed['REMARKS'].fillna('').astype(str).str.strip()
        parsed['ETS'] = parsed['ETS'].fillna('').astype(str).str.strip()
        parsed['FUEL EU'] = parsed['FUEL EU'].fillna('').astype(str).str.strip()
        parsed['VALIDITY'] = parsed['VALIDITY'].fillna('').astype(str).str.strip()

        # Filter out rows where both 20 and 40STD are NaN
        parsed = parsed[~(parsed['20'].isna() & parsed['40STD'].isna())]

        # Ensure desired columns in the output in the specified order
        desired_columns = ['POL', 'PORT', '20', '40STD', 'REMARKS', 'ETS', 'FUEL EU', 'VALIDITY']
        parsed = parsed[desired_columns]

        return parsed, remarks

    except Exception as e:
        print(f"Error in parse_msc_eur_med: {str(e)}")
        return pd.DataFrame(), [f"Error: {str(e)}"]

def parse_pil(file, month_year):
    try:
        df_pil_raw = pd.read_excel(file, sheet_name=0, header=None)
        info = []
        notes_section = False
        for idx, row in df_pil_raw.iterrows():
            for cell in row:
                cell_str = str(cell).strip()
                if any(keyword in cell_str.lower() for keyword in ['subject to', 'sub to']) and 'valid till' not in cell_str.lower():
                    notes_section = True
                if notes_section and cell_str and cell_str != 'nan':
                    info.append(cell_str)
        info = list(dict.fromkeys([i for i in info if any(keyword in i.lower() for keyword in ['subject to', 'sub to'])]))[:10]

        header_row = None
        expected_cols = ['POL', 'POD', "20'FT ($)", "40'FT ($)"]
        for idx, row in df_pil_raw.iterrows():
            row_values = row.astype(str).str.lower().str.strip()
            if all(any(col.lower() in val for val in row_values) for col in expected_cols) and not any('offered rates' in val for val in row_values):
                header_row = idx
                break

        if header_row is None:
            return pd.DataFrame(), ["Could not find header row with expected columns: 'POL', 'POD', '20'FT ($)', '40'FT ($)'. Please check the file structure."]

        df_pil = pd.read_excel(file, sheet_name=0, skiprows=header_row)
        pil_map = {
            'POL': 'POL',
            'POD': 'PORT',
            "20'FT ($)": '20',
            "40'FT ($)": '40STD',
            'ROUTINE': 'ROUTING',
            'APPX. T/TIME': 'TRANSIT TIME',
            'Validity': 'VALIDITY',
            'Remarks ': 'REMARKS'
        }

        required_cols = ['POL', 'POD', "20'FT ($)", "40'FT ($)"]
        missing_cols = [col for col in required_cols if col not in df_pil.columns]
        if missing_cols:
            return pd.DataFrame(), [f"Missing required columns: {missing_cols}. Available columns: {df_pil.columns.tolist()}"]

        base_cols = [col for col in required_cols if col in df_pil.columns]
        parsed_pil = df_pil[base_cols].rename(columns={col: pil_map[col] for col in base_cols})

        optional_cols = ['ROUTINE', 'APPX. T/TIME', 'Validity', 'Remarks ']
        for col in optional_cols:
            if col in df_pil.columns:
                parsed_pil[pil_map[col]] = df_pil[col].astype(str).str.strip().replace('nan', '')

        # Clean numeric columns and remove decimal points
        parsed_pil['20'] = clean_numeric(parsed_pil['20']).apply(
            lambda x: str(int(x)) if pd.notna(x) and x != 0 else np.nan
        )
        parsed_pil['40STD'] = clean_numeric(parsed_pil['40STD']).apply(
            lambda x: str(int(x)) if pd.notna(x) and x != 0 else np.nan
        )

        if '40HC' in parsed_pil.columns:
            parsed_pil = parsed_pil.drop(columns=['40HC'])

        parsed_pil['ROUTING'] = parsed_pil.get('ROUTING', '').astype(str).str.strip().replace('nan', '')
        parsed_pil['TRANSIT TIME'] = parsed_pil.get('TRANSIT TIME', '').astype(str).str.strip().replace('nan', '')

        parsed_pil['PORT'] = parsed_pil['PORT'].astype(str).str.strip().replace('nan', '')
        parsed_pil['POL'] = parsed_pil['POL'].astype(str).str.strip().replace('nan', '')
        parsed_pil['POL'] = parsed_pil['POL'].apply(standardize_pol)  # Standardize POL names

        parsed_pil = parsed_pil[~parsed_pil['PORT'].str.upper().str.contains('PORTS', na=False)]
        parsed_pil = parsed_pil.dropna(subset=['PORT', 'POL'])
        parsed_pil = parsed_pil[(parsed_pil['PORT'].str.strip() != '') & (parsed_pil['POL'].str.strip() != '')]

        parsed_pil = parsed_pil[~((parsed_pil['20'].isna() | (parsed_pil['20'] == 'nan')) & 
                                  (parsed_pil['40STD'].isna() | (parsed_pil['40STD'] == 'nan')))]

        # Split ports in the PORT column
        parsed_pil = split_ports(parsed_pil, port_column='PORT')

        return parsed_pil, info
    except Exception as e:
        print(f"Error parsing PIL MRG file: {str(e)}")
        return pd.DataFrame(), [f"Error parsing PIL MRG file: {str(e)}"]

def parse_arkas(file, month_year):
    try:
        df_arkas_raw = pd.read_excel(file, sheet_name=0, header=None)
        header_row = None
        for idx, row in df_arkas_raw.iterrows():
            row_values = row.astype(str).str.lower().str.strip()
            if any('pol' in val for val in row_values) and any('pod' in val for val in row_values):
                header_row = idx
                break

        if header_row is None:
            return pd.DataFrame(), ["Could not find header row with 'POL' and 'POD'."]

        df_arkas = pd.read_excel(file, sheet_name=0, skiprows=header_row)
        arkas_map = {
            'POL': 'POL',
            'POD': 'PORT',
            "t2 20'": '20',
            "t2 40'": '40STD'
        }

        required_cols = ['POL', 'POD', "t2 20'", "t2 40'"]
        missing_cols = [col for col in required_cols if col not in df_arkas.columns]
        if missing_cols:
            return pd.DataFrame(), [f"Missing required columns: {missing_cols}. Available columns: {df_arkas.columns.tolist()}"]

        base_cols = [col for col in required_cols if col in df_arkas.columns]
        parsed_arkas = df_arkas[base_cols].rename(columns={col: arkas_map[col] for col in base_cols})

        # Clean numeric columns and remove decimal points by converting to integers
        parsed_arkas['20'] = clean_numeric(parsed_arkas['20']).apply(
            lambda x: str(int(x)) if pd.notna(x) and x != 0 else np.nan
        )
        parsed_arkas['40STD'] = clean_numeric(parsed_arkas['40STD']).apply(
            lambda x: str(int(x)) if pd.notna(x) and x != 0 else np.nan
        )
        parsed_arkas['40HC'] = np.nan

        parsed_arkas['PORT'] = parsed_arkas['PORT'].astype(str).str.strip().replace('nan', '')
        parsed_arkas['POL'] = parsed_arkas['POL'].astype(str).str.strip().replace('nan', '')
        parsed_arkas['POL'] = parsed_arkas['POL'].replace({'INMUN': 'Mundra', 'INNSA': 'Nhava Sheva'})
        parsed_arkas['POL'] = parsed_arkas['POL'].apply(standardize_pol)  # Standardize POL names

        unwanted_keywords = ['PORTS', 'TOTAL', 'REMARK', 'CHARGE', 'OTHERS']
        parsed_arkas = parsed_arkas[~parsed_arkas['PORT'].str.upper().str.contains('|'.join(unwanted_keywords), na=False)]

        parsed_arkas = parsed_arkas.dropna(subset=['PORT'])
        parsed_arkas = parsed_arkas[parsed_arkas['PORT'].str.strip() != '']

        parsed_arkas = parsed_arkas[~((parsed_arkas['20'].isna() | (parsed_arkas['20'] == '0')) &
                                      (parsed_arkas['40STD'].isna() | (parsed_arkas['40STD'] == '0')))]

        parsed_arkas = parsed_arkas.dropna(subset=['POL', 'PORT'])
        parsed_arkas = parsed_arkas[(parsed_arkas['POL'].str.strip() != '') & (parsed_arkas['PORT'].str.strip() != '')]

        # Split ports in the PORT column
        parsed_arkas = split_ports(parsed_arkas, port_column='PORT')

        return parsed_arkas, []
    except Exception as e:
        print(f"Error parsing ARKAS MRG file: {str(e)}")
        return pd.DataFrame(), [f"Error parsing ARKAS MRG file: {str(e)}"]

def parse_interasia(file, month_year):
    try:
        # === STEP 1: Read Excel with pandas ===
        df_raw = pd.read_excel(file, header=6)

        # === STEP 2: Clean and map relevant columns ===
        column_map = {
            'POD': 'PORT',
            "FRT FOR 20'SD": '20',
            "FRT FOR 40'HC": '40HC',
            "FRT FOR 20'SD HAZ": '20 Haz',
            "FRT FOR 40'HC HAZ": '40 Haz',
            'ROUTING': 'ROUTING',
            'TRANSIT': 'TRANSIT TIME',
            'SERVICE': 'SERVICE',
        }
        df = df_raw[list(column_map.keys())].rename(columns=column_map)

        # Add static POL column
        df.insert(0, 'POL', 'Nhava Sheva')
        df['POL'] = df['POL'].apply(standardize_pol)  # Standardize POL names

        # === STEP 3: Dynamically extract REMARKS (was Additional Information) ===
        # Load workbook to read merged cells
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        # Find where SERVICE is
        service_col_index = df_raw.columns.get_loc('SERVICE')
        remarks_col_index = service_col_index + 1  # next column after SERVICE

        # Add REMARKS column
        df['REMARKS'] = ''

        # Build a mapping: Excel Row Number → REMARKS
        row_to_info = {}

        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            if min_col > service_col_index + 1:  # merged cell to the right of SERVICE
                merged_value = ws.cell(row=min_row, column=min_col).value
                if merged_value:
                    for r in range(min_row, max_row + 1):
                        row_to_info[r] = str(merged_value).strip()

        # Fill REMARKS based on Excel row numbers
        excel_start_row = 8  # because header is row 7, data starts at row 8
        for i in range(len(df)):
            excel_row_num = i + excel_start_row
            if excel_row_num in row_to_info:
                df.at[i, 'REMARKS'] = row_to_info[excel_row_num]
            else:
                df.at[i, 'REMARKS'] = ''  # empty if no info

        # === STEP 4: Clean Freight Columns ===
        def preserve_dollar(val):
            if pd.isna(val):
                return ''
            val_str = str(val).strip()
            if any(c.isdigit() for c in val_str) or '$' in val_str:
                cleaned = val_str.replace('USD', '').replace(' ', '')
                return f"${cleaned}" if '$' not in cleaned and any(c.isdigit() for c in cleaned) else cleaned
            return val_str

        df['20'] = df['20'].apply(preserve_dollar)
        df['40HC'] = df['40HC'].apply(preserve_dollar)
        df['20 Haz'] = df['20 Haz'].apply(preserve_dollar)
        df['40 Haz'] = df['40 Haz'].apply(preserve_dollar)

        # Merge comments across 20/40 and 20 Haz/40 Haz into REMARKS
        for index, row in df.iterrows():
            # Merge 20 and 40HC
            comment_20 = row['20'] if not any(c.isdigit() for c in str(row['20'])) else ''
            comment_40 = row['40HC'] if not any(c.isdigit() for c in str(row['40HC'])) else ''
            if comment_20 or comment_40:
                df.at[index, '20'] = comment_20 or comment_40
                df.at[index, '40HC'] = comment_20 or comment_40
                df.at[index, 'REMARKS'] = (df.at[index, 'REMARKS'] + ' ' + (comment_20 or comment_40)).strip()

            # Merge 20 Haz and 40 Haz
            comment_20_haz = row['20 Haz'] if not any(c.isdigit() for c in str(row['20 Haz'])) else ''
            comment_40_haz = row['40 Haz'] if not any(c.isdigit() for c in str(row['40 Haz'])) else ''
            if comment_20_haz or comment_40_haz:
                df.at[index, '20 Haz'] = comment_20_haz or comment_40_haz
                df.at[index, '40 Haz'] = comment_20_haz or comment_40_haz
                df.at[index, 'REMARKS'] = (df.at[index, 'REMARKS'] + ' ' + (comment_20_haz or comment_40_haz)).strip()

        df['SERVICE'] = df['SERVICE'].fillna('')

        # === STEP 5: Filter only rows with non-empty PORT ===
        df = df[
            (df['PORT'].notna() & df['PORT'].astype(str).str.strip().ne(''))
        ]

        # Split ports in the PORT column
        df = split_ports(df, port_column='PORT')

        return df, []
    except Exception as e:
        print(f"Error parsing Interasia file: {str(e)}")
        return pd.DataFrame(), []

def parse_cosco_gulf(file, month_year):
    try:
        with pdfplumber.open(file) as pdf:
            data = []
            pol_data = {pol: [] for pol in ["NHAVASHEVA", "MUNDRA", "HAZIRA"]}  # Store data per POL
            info_lines = []

            # Iterate through pages
            for page_num, page in enumerate(pdf.pages, 1):
                text = page.extract_text()
                if not text:
                    continue

                # Extract informational text
                info_match = re.search(r"(Bookings should not be placed without filing a rate.*?(?=\Z))", text, re.DOTALL)
                if info_match:
                    info_text = info_match.group(1).strip()
                    info_lines.extend([line.strip() for line in info_text.split('\n') if line.strip() and len(line.split()) > 2])

                # Split text into lines and process
                lines = [line.strip() for line in text.split('\n') if line.strip()]

                for line_idx, line in enumerate(lines):
                    if "PORT NAME" in line.upper():
                        continue
                    elif any(pol in line.upper() for pol in ["NHAVASHEVA", "MUNDRA", "HAZIRA"]):
                        continue
                    elif re.match(r'^[A-Za-z\s\-\'(),;]+\s+\d+\s+\d+\s+\d+\s+\d+\s+\d+\s+\d+$', line):
                        # Extract full port name with all punctuation and rates
                        match = re.match(r'^([A-Za-z\s\-\'(),;]+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)$', line)
                        if match:
                            port_name = match.group(1).strip()  # Full port name with punctuation
                            rate_parts = [int(x) for x in match.groups()[1:]]  # Six rate values
                            if len(rate_parts) == 6:
                                for i, pol in enumerate(["NHAVASHEVA", "MUNDRA", "HAZIRA"]):
                                    rate_20 = rate_parts[i * 2]
                                    rate_40 = rate_parts[i * 2 + 1]
                                    if rate_20 > 0 and rate_40 > 0:
                                        pol_data[pol].append({
                                            'PORT': port_name,
                                            '20': rate_20,
                                            '40STD': rate_40
                                        })

            # Combine data in POL order
            for pol in ["NHAVASHEVA", "MUNDRA", "HAZIRA"]:
                for entry in pol_data[pol]:
                    data.append({
                        'POL': pol,
                        'PORT': entry['PORT'],
                        '20': entry['20'],
                        '40STD': entry['40STD']
                    })

            # Create DataFrame
            parsed_cosco_gulf = pd.DataFrame(data) if data else pd.DataFrame(columns=['POL', 'PORT', '20', '40STD'])
            
            # Standardize column names and ensure POL is correctly formatted
            parsed_cosco_gulf['40HC'] = np.nan  # Add 40HC column for consistency
            parsed_cosco_gulf['POL'] = parsed_cosco_gulf['POL'].apply(standardize_pol)  # Standardize POL names
            parsed_cosco_gulf['PORT'] = parsed_cosco_gulf['PORT'].astype(str).str.strip()
            parsed_cosco_gulf['20'] = pd.to_numeric(parsed_cosco_gulf['20'], errors='coerce')
            parsed_cosco_gulf['40STD'] = pd.to_numeric(parsed_cosco_gulf['40STD'], errors='coerce')

            # Remove empty or invalid rows
            parsed_cosco_gulf = parsed_cosco_gulf.dropna(subset=['PORT', '20', '40STD'])
            parsed_cosco_gulf = parsed_cosco_gulf[parsed_cosco_gulf['PORT'].str.strip() != '']

            # Split ports in the PORT column
            parsed_cosco_gulf = split_ports(parsed_cosco_gulf, port_column='PORT')

            return parsed_cosco_gulf, info_lines

    except Exception as e:
        print(f"Error parsing Cosco-Gulf file: {str(e)}")
        return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC']), [f"Error parsing Cosco-Gulf file: {str(e)}"]

def parse_cosco_wcsa_cb(file, month_year):
    try:
        with pdfplumber.open(file) as pdf:
            data = []
            wcsa_remarks = []  # Isolated remarks for WCSA & CB

            # Iterate through pages
            for page_num, page in enumerate(pdf.pages, 1):
                # Extract full text for remarks and data
                text = page.extract_text()
                lines = [line.strip() for line in text.split('\n') if line.strip()]
                table_started = False
                potential_data_lines = []

                # First pass: Collect potential data and remarks
                for line in lines:
                    # Detect table start with flexible header match
                    if not table_started and any(col in line.upper() for col in ["20'", "40'", "ROUTING", "T.TIME"]):
                        table_started = True
                    elif not table_started:
                        wcsa_remarks.append(line)
                    elif table_started:
                        potential_data_lines.append(line)

                # Second pass: Process data lines, treat non-matches as remarks
                for line in potential_data_lines:
                    # Match data rows with flexible spacing
                    match = re.match(
                        r'^([A-Za-z\s,\/&-]+?)\s*(\d+)\s*(\d+)\s*([A-Za-z]+)\s*(\d+)\s*[-–]\s*(\d+)\s*days$',
                        line,
                        re.IGNORECASE
                    )
                    if match:
                        port = match.group(1).strip()
                        rate_20 = int(match.group(2))
                        rate_40 = int(match.group(3))
                        routing = match.group(4).strip()
                        t_time = f"{match.group(5)} - {match.group(6)} days"
                        if rate_20 > 0 and rate_40 > 0:
                            data.append({
                                'POL': 'Nhava Sheva',
                                'PORT': port,
                                '20': rate_20,
                                '40STD': rate_40,
                                'ROUTING': routing,
                                'TRANSIT TIME': t_time  # Renamed from T.Time (Approx)
                            })
                            data.append({
                                'POL': 'Mundra',
                                'PORT': port,
                                '20': rate_20,
                                '40STD': rate_40,
                                'ROUTING': routing,
                                'TRANSIT TIME': t_time  # Renamed from T.Time (Approx)
                            })
                    else:
                        wcsa_remarks.append(line)

            # Create DataFrame with Nhava Sheva first, then Mundra
            nhava_data = [d for d in data if d['POL'] == 'Nhava Sheva']
            mundra_data = [d for d in data if d['POL'] == 'Mundra']
            parsed_wcsa_cb = pd.DataFrame(nhava_data + mundra_data) if data else pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', 'ROUTING', 'TRANSIT TIME'])

            # Standardize column names
            parsed_wcsa_cb['40HC'] = np.nan  # Add 40HC column for consistency
            parsed_wcsa_cb['POL'] = parsed_wcsa_cb['POL'].apply(standardize_pol)  # Standardize POL names

            # Remove empty or invalid rows
            parsed_wcsa_cb = parsed_wcsa_cb.dropna(subset=['PORT', '20', '40STD'])
            parsed_wcsa_cb = parsed_wcsa_cb[parsed_wcsa_cb['PORT'].str.strip() != '']

            # Split ports in the PORT column
            parsed_wcsa_cb = split_ports(parsed_wcsa_cb, port_column='PORT')

            return parsed_wcsa_cb, wcsa_remarks

    except Exception as e:
        print(f"Error parsing Cosco-WCSA & CB file: {str(e)}")
        return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'ROUTING', 'TRANSIT TIME']), [f"Error parsing Cosco-WCSA & CB file: {str(e)}"]

def parse_cosco_africa(file, month_year):
    try:
        records = []
        raw_lines = []
        current_section = ""
        pipavav_allowed = False
        last_matched_line_index = -1
        port_sequence = []
        port_seen = set()

        with pdfplumber.open(file) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue
                raw_lines.extend([line.strip() for line in text.split('\n') if line.strip()])

        for idx, line in enumerate(raw_lines):
            # Detect section headers
            if re.search(r'WEST AFRICA.*SOUTH WEST AFRICA', line, re.I):
                current_section = line.strip()
                pipavav_allowed = True
                continue
            elif re.search(r'\bEAST AFRICA\b', line, re.I):
                current_section = line.strip()
                pipavav_allowed = False
                continue
            elif re.search(r'\bSOUTH AFRICA\b', line, re.I):
                current_section = line.strip()
                pipavav_allowed = False
                continue

            # Match port + rates
            match = re.match(r'^(.+?)\s+(\d{3,5})\s+(\d{3,5})$', line)
            if match:
                port = match.group(1).strip()
                rate_20 = int(match.group(2))
                rate_40 = int(match.group(3))

                if port not in port_seen:
                    port_sequence.append(port)
                    port_seen.add(port)
                last_matched_line_index = idx

                for pol in ['Nhava Sheva', 'Mundra']:
                    records.append({
                        'POL': pol,
                        'PORT': port,
                        '20': rate_20,
                        '40STD': rate_40,
                        'REMARKS': current_section
                    })
                if pipavav_allowed:
                    records.append({
                        'POL': 'Pipavav',
                        'PORT': port,
                        '20': rate_20,
                        '40STD': rate_40,
                        'REMARKS': current_section
                    })

        # Collect all remaining lines as post-table remarks
        full_post_table_remarks = raw_lines[last_matched_line_index + 1:]

        # Create DataFrame
        parsed_cosco_africa = pd.DataFrame(records)

        # Standardize column names and ensure POL is correctly formatted
        parsed_cosco_africa['40HC'] = np.nan  # Add 40HC column for consistency
        parsed_cosco_africa['POL'] = parsed_cosco_africa['POL'].apply(standardize_pol)  # Standardize POL names

        # Remove empty or invalid rows
        parsed_cosco_africa = parsed_cosco_africa.dropna(subset=['PORT', '20', '40STD'])
        parsed_cosco_africa = parsed_cosco_africa[parsed_cosco_africa['PORT'].str.strip() != '']

        # Split ports in the PORT column
        parsed_cosco_africa = split_ports(parsed_cosco_africa, port_column='PORT')

        return parsed_cosco_africa, full_post_table_remarks

    except Exception as e:
        print(f"Error parsing Cosco-Africa file: {str(e)}")
        return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'REMARKS']), [f"Error parsing Cosco-Africa file: {str(e)}"]

def parse_cosco_fareast(file, month_year):
    try:
        import re
        import pdfplumber
        pols = ["Nhava Sheva", "Mundra", "Pipavav"]
        all_rows = []
        additional_info = []

        # Pattern: PORT + 3 values (number or 'CASE BY CASE')
        pattern = re.compile(
            r"([A-Za-z\s\-/().,'&]+?)\s+(CASE BY CASE|\d+(?:\.\d+)?)\s*(CASE BY CASE|\d+(?:\.\d+)?)\s*(CASE BY CASE|\d+(?:\.\d+)?)",
            re.IGNORECASE
        )

        def parse_value(v):
            v = v.upper().strip()
            return v if "CASE" in v else float(v)

        with pdfplumber.open(file) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue
                lines = [line.strip() for line in text.split('\n') if line.strip()]

                for line in lines:
                    if line.startswith("*") or any(keyword in line.lower() for keyword in [
                        "rates are", "currently", "note", "booking", "surcharge",
                        "valid", "no service", "not accepting", "weight limitation"]):
                        additional_info.append(line)
                        continue

                    # 🛠️ Fix glued tokens like 'CASE BY CASECASE BY CASE'
                    line = re.sub(r'(CASE BY CASE)(CASE BY CASE)', r'\1 \2', line, flags=re.IGNORECASE)

                    matches = pattern.findall(line)
                    if not matches:
                        continue

                    if len(matches) == 3:
                        # Assign each match to Nhava Sheva, Mundra, Pipavav
                        for i, match in enumerate(matches):
                            port, r20, r40std, r40hc = match
                            pol = pols[i]
                            try:
                                r20_val = parse_value(r20)
                                r40std_val = parse_value(r40std)
                                r40hc_val = parse_value(r40hc)

                                all_rows.append({
                                    'POL': pol,
                                    'PORT': port.strip(),
                                    '20': r20_val,
                                    '40STD': r40std_val,
                                    '40HC': r40hc_val,
                                    'REMARKS': ''
                                })
                            except Exception as e:
                                print(f"Error parsing CASE row: {e}")
                    else:
                        # Fallback: assign POLs in rotating order
                        for idx, match in enumerate(matches):
                            port, r20, r40std, r40hc = match
                            pol = pols[idx % 3]
                            try:
                                r20_val = parse_value(r20)
                                r40std_val = parse_value(r40std)
                                r40hc_val = parse_value(r40hc)

                                all_rows.append({
                                    'POL': pol,
                                    'PORT': port.strip(),
                                    '20': r20_val,
                                    '40STD': r40std_val,
                                    '40HC': r40hc_val,
                                    'REMARKS': ''
                                })
                            except Exception:
                                continue

        df = pd.DataFrame(all_rows)

        if not df.empty:
            df['POL'] = df['POL'].apply(standardize_pol)
            df['PORT'] = df['PORT'].astype(str).str.strip()
            df = df[df['PORT'].notna() & (df['PORT'].str.strip() != '')]

            df['REMARKS'] = df['REMARKS'].astype(str).str.strip()
            df['REMARKS'] = df['REMARKS'].replace('', np.nan)

            # Convert numeric values to integers (removing decimal points) and store as strings
            for col in ['20', '40STD', '40HC']:
                df[col] = df[col].apply(
                    lambda x: str(int(x)) if isinstance(x, float) and pd.notna(x) and x != 0 else x
                )

            df['POL_ORDER'] = df['POL'].map({'Nhava Sheva': 1, 'Mundra': 2, 'Pipavav': 3})
            df = df.sort_values(by=['POL_ORDER', 'PORT']).drop(columns=['POL_ORDER'])
            df = df.dropna(subset=['20', '40STD', '40HC'], how='all')

        # Split ports in the PORT column
        df = split_ports(df, port_column='PORT')

        additional_info = list(dict.fromkeys(additional_info))
        return df, additional_info

    except Exception as e:
        return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'REMARKS']), [f"Error parsing Cosco-Fareast file: {str(e)}"]


from openpyxl import load_workbook
import traceback
import re

def parse_zim(file, month_year):
    try:
        xl = pd.ExcelFile(file)
        sheets = xl.sheet_names
        data_frames = []
        all_remarks = []

        def clean_numeric_series(col):
            def clean_value(val):
                if pd.isna(val) or val is None or str(val).strip() == '':
                    return np.nan
                if isinstance(val, (int, float)):
                    return float(val)
                val_str = str(val).strip()
                val_clean = val_str.replace(',', '').replace('EUR', '').replace('USD', '').replace('$', '').replace('€', '').replace(' ', '').strip()
                if not val_clean.replace('.', '').replace('-', '').isdigit():
                    print(f"Invalid value in clean_numeric_series: '{val_str}'")
                    return np.nan
                try:
                    return float(val_clean)
                except (ValueError, TypeError) as e:
                    print(f"Error converting value '{val_str}' to float: {str(e)}")
                    return np.nan
            return col.apply(clean_value)

        def standardize_pol(pol):
            if pd.isna(pol):
                return ''
            pol = str(pol).strip().upper()
            mapping = {
                'INNHV': 'Nhava Sheva',
                'NHAVA SHEVA': 'Nhava Sheva',
                'NHV/RQ': 'Nhava Sheva',
                'NHAVA': 'Nhava Sheva',
                'NSA': 'Nhava Sheva',
                'NS': 'Nhava Sheva',
                'MUNDRA': 'Mundra',
                'PIPAVAV': 'Pipavav',
                'TUTICORIN': 'Tut',
                'TUT': 'Tut',
                'KATTUPALLI': 'Ktp',
                'KOLKATA': 'Ccu',
                'CCU': 'Ccu',
                'WRG': 'Wrg',
                'INMUN': 'Mundra',
                'INNSA': 'Nhava Sheva',
                'HAZIRA': 'Hazira',
                'HZ': 'Hazira',
                'INHZA': 'Hazira'
            }
            return mapping.get(pol, pol.title())

        def is_cell_strikethrough(cell):
            """Check if a cell has strikethrough formatting."""
            if cell.value is None:
                return False
            font = cell.font
            return font.strike if font else False
        
        def format_rates_with_dollar(df):
            for col in ['20', '40STD', '40HC']:
                if col in df.columns:
                    df[col] = df[col].apply(
                        lambda x: f"${int(x)}" if pd.notna(x) and isinstance(x, (int, float)) and x == int(x)
                        else f"${x}" if pd.notna(x) else ""
                    )
            return df

        def process_turkey_med_sheet(sheet_name):
            remarks = []
            try:
                wb = load_workbook(file, data_only=True)
                ws = wb[sheet_name]
                records = []
                parsed_ports = set()

                for idx, row in enumerate(ws.iter_rows(min_row=3), start=3):
                    pod_cell = row[0]
                    if not pod_cell or pod_cell.value is None or is_cell_strikethrough(pod_cell):
                        continue

                    pod = str(pod_cell.value).strip()
                    if not pod or pod.upper().startswith("SUBJECT") or pod.upper().startswith("POL"):
                        break  # End of data section
                    if len(pod.split()) > 3:
                        remarks.append(pod)
                        continue

                    parsed_ports.add(pod.title())

                    def get_cleaned(idx):
                        if idx >= len(row):
                            return None
                        cell = row[idx]
                        return None if is_cell_strikethrough(cell) else cell.value

                    nhv_20 = clean_numeric_series(pd.Series([get_cleaned(1)]))[0]
                    nhv_40 = clean_numeric_series(pd.Series([get_cleaned(2)]))[0]
                    hzr_20 = clean_numeric_series(pd.Series([get_cleaned(3)]))[0]
                    hzr_40 = clean_numeric_series(pd.Series([get_cleaned(4)]))[0]

                    if pd.notna(nhv_20) or pd.notna(nhv_40):
                        records.append({
                            'POL': 'Nhava Sheva',
                            'PORT': pod.title(),
                            '20': nhv_20,
                            '40STD': None,
                            '40HC': nhv_40,
                            'Remarks': ''
                        })
                    if pd.notna(hzr_20) or pd.notna(hzr_40):
                        records.append({
                            'POL': 'Hazira',
                            'PORT': pod.title(),
                            '20': hzr_20,
                            '40STD': None,
                            '40HC': hzr_40,
                            'Remarks': ''
                        })

                # ✅ Collect Additional Information (true remarks only)
                seen_lines = set()
                for row in ws.iter_rows(min_row=1, max_row=80):
                    line = " ".join(str(cell.value).strip() for cell in row if cell.value and str(cell.value).strip())
                    if not line or line in seen_lines:
                        continue
                    seen_lines.add(line)

                    # Skip lines that start with a parsed POD
                    if any(line.upper().startswith(pod.upper()) for pod in parsed_ports):
                        continue

                    # Skip lines that look like tabular data (≥3 numeric tokens)
                    numeric_token_count = sum(1 for token in line.split() if token.replace('.', '', 1).isdigit())
                    if numeric_token_count >= 3:
                        continue

                    # Keep meaningful remarks only
                    if len(line.split()) > 3:
                        remarks.append(line)

                df_final = pd.DataFrame(records)
                if df_final.empty:
                    print(f"No valid records parsed from sheet '{sheet_name}'")
                    return df_final, remarks

                pol_priority = {'Nhava Sheva': 1, 'Hazira': 2}
                df_final['POL_ORDER'] = df_final['POL'].map(pol_priority)
                df_final = df_final.sort_values(by='POL_ORDER').drop(columns='POL_ORDER')
                df_final = df_final[['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']]
                print(f"Parsed {len(df_final)} records from sheet '{sheet_name}', collected {len(remarks)} remarks")
                return df_final, list(dict.fromkeys(remarks))  # unique remarks

            except Exception as e:
                print(f"Error in process_turkey_med_sheet for sheet '{sheet_name}': {str(e)}")
                print(f"Full traceback: {traceback.format_exc()}")
                return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), [f"Error: {str(e)}"]


        def fallback_parse_turkey_med(file, sheet_name, remarks):
            print(f"Attempting fallback parsing for sheet '{sheet_name}' using pandas")
            try:
                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None)
                wb = load_workbook(file, data_only=True)
                ws_data = wb[sheet_name]
                print(f"Read {len(df_raw)} rows from sheet '{sheet_name}' with pandas")
                
                records = []
                header_row_idx = None
                col_map = {}

                # Find header row
                for row_idx, row in enumerate(ws_data.iter_rows(min_row=1, max_row=20), start=1):
                    row_vals = [str(cell.value).strip().upper() for cell in row if cell.value not in (None, '') and not is_cell_strikethrough(cell)]
                    row_str = ' '.join(row_vals)
                    if any(x in row_str for x in ['NHAVA', 'NS', 'NSA', 'NHV', 'INNHV', 'HAZIRA', 'HZ', 'INHZA', 'POD', 'PORT', 'DEST', 'DESTINATION', 'DISCHARGE', '20', '40', 'HC']):
                        header_row_idx = row_idx - 1
                        header_row = row
                        break

                if header_row_idx is None:
                    print(f"Fallback: No header row found in sheet '{sheet_name}' after searching 20 rows")
                    for row_idx, row in enumerate(ws_data.iter_rows(min_row=1, max_row=50), start=1):
                        print(f"Row {row_idx}: {[str(cell.value) for cell in row]}")
                    return data_driven_parse_turkey_med(df_raw, sheet_name, remarks)

                print(f"Fallback: Found header row {header_row_idx+1} in '{sheet_name}': {[str(cell.value) for cell in header_row]}")
                row_vals = [str(cell.value).strip().upper() for cell in header_row if cell.value and not is_cell_strikethrough(cell)]
                pol_cols = {}
                if any(pol in ' '.join(row_vals) for pol in ['INNHV', 'INHZA', 'NHAVA', 'HAZIRA']):
                    if header_row_idx + 1 < len(df_raw):
                        next_row = [cell for cell in ws_data[header_row_idx + 1]]
                        next_row_vals = [str(cell.value).strip().upper() for cell in next_row if cell.value and not is_cell_strikethrough(cell)]
                        print(f"Fallback: Checking next row {header_row_idx+2} for container types: {[str(cell.value) for cell in next_row]}")
                        if any(ct in ' '.join(next_row_vals) for ct in ['20', '40', 'HC']):
                            pol_row = header_row
                            container_row = next_row
                            header_row_idx += 1
                            for col_idx, (pol_cell, cont_cell) in enumerate(zip(pol_row, container_row)):
                                if is_cell_strikethrough(pol_cell) or is_cell_strikethrough(cont_cell):
                                    continue
                                pol_val = str(pol_cell.value).strip().upper() if pol_cell.value else ''
                                cont_val = str(cont_cell.value).strip().upper() if cont_cell.value else ''
                                if 'INNHV' in pol_val or 'NHAVA' in pol_val:
                                    if '20' in cont_val:
                                        col_map['nhv_20'] = col_idx
                                        pol_cols[col_idx] = 'Nhava Sheva'
                                    elif '40' in cont_val or 'HC' in cont_val:
                                        col_map['nhv_40'] = col_idx
                                        pol_cols[col_idx] = 'Nhava Sheva'
                                elif 'INHZA' in pol_val or 'HAZIRA' in pol_val:
                                    if '20' in cont_val:
                                        col_map['hzr_20'] = col_idx
                                        pol_cols[col_idx] = 'Hazira'
                                    elif '40' in cont_val or 'HC' in cont_val:
                                        col_map['hzr_40'] = col_idx
                                        pol_cols[col_idx] = 'Hazira'
                            print(f"Fallback: Updated column mapping after container row: {col_map}")

                for col_idx, cell in enumerate(header_row):
                    if is_cell_strikethrough(cell):
                        continue
                    val = str(cell.value).strip().upper() if cell.value else ''
                    if any(p in val for p in ['POD', 'PORT', 'DEST', 'DESTINATION', 'DISCHARGE', 'DISCHARGE PORT', 'DEST. PORT']):
                        col_map['pod'] = col_idx
                        break

                if 'pod' not in col_map:
                    for row_idx, row in enumerate(ws_data.iter_rows(min_row=header_row_idx+1, max_row=header_row_idx+10), start=header_row_idx+1):
                        for col_idx, cell in enumerate(row):
                            if is_cell_strikethrough(cell):
                                continue
                            val = str(cell.value).strip() if cell.value else ''
                            if not val:
                                continue
                            is_numeric = bool(re.match(r'^-?\d+(\.\d+)?$', val.replace(',', '').replace(' ', '')))
                            is_pol = any(pol in val.upper() for pol in ['INNHV', 'INHZA', 'NHAVA', 'HAZIRA', 'NS', 'NSA', 'NHV', 'HZ'])
                            is_container = any(ct in val.upper() for ct in ['20', '40', 'HC'])
                            if not is_numeric and not is_pol and not is_container and len(val) > 1:
                                col_map['pod'] = col_idx
                                print(f"Fallback: Inferred POD column {col_idx} based on value '{val}'")
                                break
                        if 'pod' in col_map:
                            break

                if 'pod' not in col_map:
                    print(f"Fallback: No POD column identified in sheet '{sheet_name}'")
                    for idx, row in df_raw.head(50).iterrows():
                        print(f"Row {idx+1}: {[str(val) for val in row]}")
                    return data_driven_parse_turkey_med(df_raw, sheet_name, remarks)

                print(f"Fallback column mapping for '{sheet_name}': {col_map}")

                for row_idx, row in enumerate(ws_data.iter_rows(min_row=header_row_idx+1), start=header_row_idx+1):
                    pod_cell = row[col_map['pod']] if col_map['pod'] < len(row) else None
                    if pod_cell and is_cell_strikethrough(pod_cell):
                        print(f"Fallback: Skipping row {row_idx+1} in sheet '{sheet_name}': POD '{pod_cell.value}' is struck through")
                        continue
                    pod = str(pod_cell.value).strip() if pod_cell and pod_cell.value not in (None, '') else ''
                    if not pod:
                        print(f"Fallback: Skipping row {row_idx+1} in sheet '{sheet_name}': Empty POD")
                        continue
                    if len(pod.split()) > 3:
                        remarks.append(pod)
                        continue

                    nhv_20 = row[col_map['nhv_20']].value if 'nhv_20' in col_map and col_map['nhv_20'] < len(row) and not is_cell_strikethrough(row[col_map['nhv_20']]) else None
                    nhv_40 = row[col_map['nhv_40']].value if 'nhv_40' in col_map and col_map['nhv_40'] < len(row) and not is_cell_strikethrough(row[col_map['nhv_40']]) else None
                    hzr_20 = row[col_map['hzr_20']].value if 'hzr_20' in col_map and col_map['hzr_20'] < len(row) and not is_cell_strikethrough(row[col_map['hzr_20']]) else None
                    hzr_40 = row[col_map['hzr_40']].value if 'hzr_40' in col_map and col_map['hzr_40'] < len(row) and not is_cell_strikethrough(row[col_map['hzr_40']]) else None

                    nhv_20_clean = clean_numeric_series(pd.Series([nhv_20]))[0]
                    nhv_40_clean = clean_numeric_series(pd.Series([nhv_40]))[0]
                    hzr_20_clean = clean_numeric_series(pd.Series([hzr_20]))[0]
                    hzr_40_clean = clean_numeric_series(pd.Series([hzr_40]))[0]

                    row_data = {
                        'Nhava Sheva 20': nhv_20,
                        'Nhava Sheva 40HC': nhv_40,
                        'Hazira 20': hzr_20,
                        'Hazira 40HC': hzr_40
                    }
                    print(f"Fallback: Row {row_idx+1} in '{sheet_name}': POD={pod}, Rates={row_data}")

                    if pd.notna(nhv_20_clean) or pd.notna(nhv_40_clean):
                        records.append({
                            'POL': 'Nhava Sheva',
                            'PORT': pod.title(),
                            '20': nhv_20_clean,
                            '40STD': None,
                            '40HC': nhv_40_clean,
                            'Remarks': ''
                        })
                    if pd.notna(hzr_20_clean) or pd.notna(hzr_40_clean):
                        records.append({
                            'POL': 'Hazira',
                            'PORT': pod.title(),
                            '20': hzr_20_clean,
                            '40STD': None,
                            '40HC': hzr_40_clean,
                            'Remarks': ''
                        })
                    if not (pd.notna(nhv_20_clean) or pd.notna(nhv_40_clean) or pd.notna(hzr_20_clean) or pd.notna(hzr_40_clean)):
                        print(f"Fallback: Skipping row {row_idx+1} in sheet '{sheet_name}': No valid rates (Nhava Sheva: 20={nhv_20}, 40HC={nhv_40}, Hazira: 20={hzr_20}, 40HC={hzr_40})")

                df_final = pd.DataFrame(records)
                if df_final.empty:
                    print(f"Fallback: No valid records parsed from sheet '{sheet_name}'")
                    return df_final, remarks

                pol_priority = {'Nhava Sheva': 1, 'Hazira': 2}
                df_final['POL_ORDER'] = df_final['POL'].map(pol_priority)
                df_final = df_final.sort_values(by='POL_ORDER').drop(columns='POL_ORDER')
                df_final = df_final[['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']]
                print(f"Fallback: Parsed {len(df_final)} records from sheet '{sheet_name}', collected {len(remarks)} remarks")
                return df_final, remarks
            except Exception as e:
                print(f"Fallback parsing failed for sheet '{sheet_name}': {str(e)}")
                print(f"Full traceback: {traceback.format_exc()}")
                return data_driven_parse_turkey_med(df_raw, sheet_name, remarks)

        def data_driven_parse_turkey_med(df_raw, sheet_name, remarks):
            print(f"Attempting data-driven parsing for sheet '{sheet_name}'")
            records = []
            col_map = {}

            for col_idx in range(min(df_raw.shape[1], 10)):
                column = df_raw.iloc[:, col_idx].dropna().astype(str).str.strip()
                non_numeric_count = 0
                total_count = 0
                for val in column:
                    val_upper = val.upper()
                    is_numeric = bool(re.match(r'^-?\d+(\.\d+)?$', val.replace(',', '').replace(' ', '')))
                    is_pol = any(pol in val_upper for pol in ['INNHV', 'INHZA', 'NHAVA', 'HAZIRA', 'NS', 'NSA', 'NHV', 'HZ'])
                    is_container = any(ct in val_upper for ct in ['20', '40', 'HC'])
                    if not is_numeric and not is_pol and not is_container and len(val) > 1:
                        non_numeric_count += 1
                    total_count += 1
                if total_count > 0 and (non_numeric_count / total_count) > 0.5:
                    col_map['pod'] = col_idx
                    print(f"Data-driven: Inferred POD column {col_idx} based on non-numeric values")
                    break

            if 'pod' not in col_map:
                print(f"Data-driven: No POD column identified in sheet '{sheet_name}'")
                for idx, row in df_raw.head(50).iterrows():
                    print(f"Row {idx+1}: {[str(val) for val in row]}")
                return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), remarks + [f"Data-driven parsing failed: No POD column in '{sheet_name}'"]

            pol_cols = {}
            for col_idx in range(df_raw.shape[1]):
                column = df_raw.iloc[:, col_idx].dropna().astype(str).str.strip().str.upper()
                if any('NHAVA' in val or 'NS' in val or 'NSA' in val or 'INNHV' in val for val in column):
                    for offset in range(1, 3):
                        if col_idx + offset >= df_raw.shape[1]:
                            break
                        next_col = df_raw.iloc[:, col_idx + offset].dropna().astype(str).str.strip().str.upper()
                        is_rate_col = any(bool(re.match(r'^-?\d+(\.\d+)?$', val.replace(',', ''))) for val in next_col)
                        if is_rate_col:
                            if any('20' in val for val in next_col):
                                col_map['nhv_20'] = col_idx + offset
                                pol_cols[col_idx + offset] = 'Nhava Sheva'
                            elif any('40' in val or 'HC' in val for val in next_col):
                                col_map['nhv_40'] = col_idx + offset
                                pol_cols[col_idx + offset] = 'Nhava Sheva'
                elif any('HAZIRA' in val or 'HZ' in val or 'INHZA' in val for val in column):
                    for offset in range(1, 3):
                        if col_idx + offset >= df_raw.shape[1]:
                            break
                        next_col = df_raw.iloc[:, col_idx + offset].dropna().astype(str).str.strip().str.upper()
                        is_rate_col = any(bool(re.match(r'^-?\d+(\.\d+)?$', val.replace(',', ''))) for val in next_col)
                        if is_rate_col:
                            if any('20' in val for val in next_col):
                                col_map['hzr_20'] = col_idx + offset
                                pol_cols[col_idx + offset] = 'Hazira'
                            elif any('40' in val or 'HC' in val for val in next_col):
                                col_map['hzr_40'] = col_idx + offset
                                pol_cols[col_idx + offset] = 'Hazira'

            print(f"Data-driven column mapping for '{sheet_name}': {col_map}")

            for row_idx, row in df_raw.iterrows():
                pod = str(row[col_map['pod']]).strip() if col_map['pod'] < len(row) and pd.notna(row[col_map['pod']]) else ''
                is_numeric = bool(re.match(r'^-?\d+(\.\d+)?$', pod.replace(',', '').replace(' ', '')))
                is_pol = any(pol in pod.upper() for pol in ['INNHV', 'INHZA', 'NHAVA', 'HAZIRA', 'NS', 'NSA', 'NHV', 'HZ'])
                is_container = any(ct in pod.upper() for ct in ['20', '40', 'HC'])
                if not pod or is_numeric or is_pol or is_container:
                    print(f"Data-driven: Skipping row {row_idx+1} in sheet '{sheet_name}': Invalid or empty POD '{pod}'")
                    continue
                if len(pod.split()) > 3:
                    remarks.append(pod)
                    continue

                nhv_20 = row[col_map['nhv_20']] if 'nhv_20' in col_map and col_map['nhv_20'] < len(row) else None
                nhv_40 = row[col_map['nhv_40']] if 'nhv_40' in col_map and col_map['nhv_40'] < len(row) else None
                hzr_20 = row[col_map['hzr_20']] if 'hzr_20' in col_map and col_map['hzr_20'] < len(row) else None
                hzr_40 = row[col_map['hzr_40']] if 'hzr_40' in col_map and col_map['hzr_40'] < len(row) else None

                nhv_20_clean = clean_numeric_series(pd.Series([nhv_20]))[0]
                nhv_40_clean = clean_numeric_series(pd.Series([nhv_40]))[0]
                hzr_20_clean = clean_numeric_series(pd.Series([hzr_20]))[0]
                hzr_40_clean = clean_numeric_series(pd.Series([hzr_40]))[0]

                row_data = {
                    'Nhava Sheva 20': nhv_20,
                    'Nhava Sheva 40HC': nhv_40,
                    'Hazira 20': hzr_20,
                    'Hazira 40HC': hzr_40
                }
                print(f"Data-driven: Row {row_idx+1} in '{sheet_name}': POD={pod}, Rates={row_data}")

                if pd.notna(nhv_20_clean) or pd.notna(nhv_40_clean):
                    records.append({
                        'POL': 'Nhava Sheva',
                        'PORT': pod.title(),
                        '20': nhv_20_clean,
                        '40STD': None,
                        '40HC': nhv_40_clean,
                        'Remarks': ''
                    })
                if pd.notna(hzr_20_clean) or pd.notna(hzr_40_clean):
                    records.append({
                        'POL': 'Hazira',
                        'PORT': pod.title(),
                        '20': hzr_20_clean,
                        '40STD': None,
                        '40HC': hzr_40_clean,
                        'Remarks': ''
                    })
                if not (pd.notna(nhv_20_clean) or pd.notna(nhv_40_clean) or pd.notna(hzr_20_clean) or pd.notna(hzr_40_clean)):
                    print(f"Data-driven: Skipping row {row_idx+1} in sheet '{sheet_name}': No valid rates (Nhava Sheva: 20={nhv_20}, 40HC={nhv_40}, Hazira: 20={hzr_20}, 40HC={hzr_40})")

            df_final = pd.DataFrame(records)
            if df_final.empty:
                print(f"Data-driven: No valid records parsed from sheet '{sheet_name}'")
                return df_final, remarks

            pol_priority = {'Nhava Sheva': 1, 'Hazira': 2}
            df_final['POL_ORDER'] = df_final['POL'].map(pol_priority)
            df_final = df_final.sort_values(by='POL_ORDER').drop(columns='POL_ORDER')
            df_final = df_final[['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']]
            print(f"Data-driven: Parsed {len(df_final)} records from sheet '{sheet_name}', collected {len(remarks)} remarks")
            return df_final, remarks

        def process_generic_sheet(sheet_name, pol_name):
            remarks = []
            try:
                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None)
                print(f"Processing sheet '{sheet_name}' with {len(df_raw)} rows")
                df_raw = df_raw.iloc[4:].copy()
                df_raw.columns = ['PORT', '20', '40HC', 'Remarks', 'DG_20', 'DG_40']
                data_rows = []
                for idx, row in df_raw.iterrows():
                    port = str(row['PORT']).strip() if pd.notna(row['PORT']) else ''
                    if not port:
                        continue
                    rate_20 = row['20']
                    rate_40hc = row['40HC']
                    if pd.isna(rate_20) and pd.isna(rate_40hc):
                        print(f"Skipping row {idx} in sheet '{sheet_name}': Both '20' and '40HC' are NaN")
                        continue
                    if len(port.split()) <= 3:
                        data_rows.append(row)
                    elif len(port.split()) > 3:
                        remarks.append(port.strip())
                if not data_rows:
                    print(f"No valid data rows in sheet '{sheet_name}'")
                    return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), remarks
                df_data = pd.DataFrame(data_rows)
                df_data['POL'] = 'Hazira' if 'hazira' in sheet_name.lower() else 'Nhava Sheva'
                df_data['PORT'] = df_data['PORT'].astype(str).str.strip()
                print(f"Sheet '{sheet_name}' rate columns sample: {df_data[['20', '40HC']].head().to_dict()}")
                df_data['20'] = clean_numeric_series(df_data['20'])
                df_data['40HC'] = clean_numeric_series(df_data['40HC'])
                df_data['Remarks'] = df_data['Remarks'].astype(str).str.strip().replace('nan', '')
                df_data['40STD'] = np.nan
                parsed_df = df_data[['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']]
                parsed_df = parsed_df.dropna(subset=['PORT'])
                parsed_df = parsed_df[parsed_df['PORT'].str.strip() != '']
                parsed_df = parsed_df[~(parsed_df['20'].isna() & parsed_df['40HC'].isna())]
                print(f"Parsed {len(parsed_df)} rows from sheet '{sheet_name}', collected {len(remarks)} remarks")
                return parsed_df, remarks
            except Exception as e:
                print(f"Error in process_generic_sheet for sheet '{sheet_name}': {str(e)}")
                return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), remarks + [f"Error in '{sheet_name}': {str(e)}"]

        def process_australia_sheet(sheet_name):
            remarks = []
            try:
                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None)
                records = []
                current_pod_group = None
                reading_block = False
                for idx, row in df_raw.iterrows():
                    row_vals = [cell for cell in row]
                    if pd.notna(row_vals[0]) and isinstance(row_vals[1], str):
                        if "SYDNEY" in row_vals[1].upper():
                            current_pod_group = "SYDNEY/MELBOURNE/BRISBANE"
                            reading_block = False
                            continue
                        elif "AKL" in row_vals[1].upper():
                            current_pod_group = "AKL/TAU"
                            reading_block = False
                            continue
                    if row_vals[0] == "POL":
                        reading_block = True
                        continue
                    pol_raw = str(row_vals[0]).strip() if pd.notna(row_vals[0]) else ""
                    if reading_block and pol_raw in ["Nhava Sheva", "Mundra", "Hazira"]:
                        rate_20 = clean_numeric_series(pd.Series([row_vals[1]]))[0] if pd.notna(row_vals[1]) else None
                        rate_40 = clean_numeric_series(pd.Series([row_vals[2]]))[0] if pd.notna(row_vals[2]) else None
                        if pd.isna(rate_20) and pd.isna(rate_40):
                            continue
                        for pod in current_pod_group.split("/"):
                            pod_clean = pod.strip().title()
                            records.append({
                                'POL': pol_raw,
                                'PORT': pod_clean,
                                '20': rate_20,
                                '40HC': rate_40,
                                '40STD': None,
                                'Remarks': ''
                            })
                        continue
                    if row_vals[0] not in ["Nhava Sheva", "Mundra", "Hazira"] and any(isinstance(val, str) and len(val.strip()) > 5 for val in row_vals if pd.notna(val)):
                        joined = " ".join(str(val).strip() for val in row_vals if pd.notna(val))
                        if joined not in remarks:
                            remarks.append(joined)
                df_final = pd.DataFrame(records)
                if df_final.empty:
                    print(f"No valid records parsed from sheet '{sheet_name}'")
                    return df_final, remarks
                df_final = df_final[['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']]
                print(f"Parsed {len(df_final)} records from sheet '{sheet_name}', collected {len(remarks)} remarks")
                return df_final, remarks
            except Exception as e:
                print(f"Error in process_australia_sheet for sheet '{sheet_name}': {str(e)}")
                return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), remarks + [f"Error in '{sheet_name}': {str(e)}"]

        def process_ecsa_sheet(sheet_name):
            remarks = []
            try:
                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None)
                pol_row_idx = 0
                pol_row = df_raw.iloc[pol_row_idx].dropna().astype(str).str.upper()
                pols = [standardize_pol(p.strip()) for p in pol_row.iloc[0].split('/')] if not pol_row.empty else ["Nhava Sheva", "Mundra"]
                data_start_row = 1
                data_end_row = None
                for idx in range(data_start_row, len(df_raw)):
                    port = str(df_raw.iloc[idx, 0]).strip() if pd.notna(df_raw.iloc[idx, 0]) else ''
                    if not port or any(keyword in port.upper() for keyword in ["ABOVE", "SUB", "DG", "VALIDITY", "ROUTE", "SUBJECT", "INCU", "INKT"]):
                        data_end_row = idx
                        break
                if data_end_row is None:
                    data_end_row = len(df_raw)
                port_data = []
                for idx in range(data_start_row, data_end_row):
                    port = str(df_raw.iloc[idx, 0]).strip() if pd.notna(df_raw.iloc[idx, 0]) else ''
                    if not port:
                        continue
                    rate_20 = df_raw.iloc[idx, 1] if pd.notna(df_raw.iloc[idx, 1]) else None
                    rate_40 = df_raw.iloc[idx, 2] if pd.notna(df_raw.iloc[idx, 2]) else None
                    rate_20_clean = clean_numeric_series(pd.Series([rate_20]))[0]
                    rate_40_clean = clean_numeric_series(pd.Series([rate_40]))[0]
                    if pd.isna(rate_20_clean) and pd.isna(rate_40_clean):
                        continue
                    port_data.append({
                        'PORT': port.title(),
                        '20': rate_20_clean,
                        '40HC': rate_40_clean,
                        'Remarks': ''
                    })
                records = []
                for data in port_data:
                    records.append({
                        'POL': 'Nhava Sheva',
                        'PORT': data['PORT'],
                        '20': data['20'],
                        '40HC': data['40HC'],
                        'Remarks': data['Remarks']
                    })
                for data in port_data:
                    records.append({
                        'POL': 'Mundra',
                        'PORT': data['PORT'],
                        '20': data['20'],
                        '40HC': data['40HC'],
                        'Remarks': data['Remarks']
                    })
                for idx in range(data_end_row, len(df_raw)):
                    row = df_raw.iloc[idx]
                    for cell in row:
                        if pd.notna(cell) and str(cell).strip():
                            remarks.append(str(cell).strip())
                df_final = pd.DataFrame(records)
                if df_final.empty:
                    print(f"No valid records parsed from sheet '{sheet_name}'")
                    return df_final, remarks
                df_final['40STD'] = np.nan
                df_final = df_final[['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']]
                print(f"Parsed {len(df_final)} records from sheet '{sheet_name}', collected {len(remarks)} remarks")
                return df_final, remarks
            except Exception as e:
                print(f"Error in process_ecsa_sheet for sheet '{sheet_name}': {str(e)}")
                return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), remarks + [f"Error in '{sheet_name}': {str(e)}"]

        def process_gulf_sheet(sheet_name):
            remarks = []
            try:
                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None)
                port_row = df_raw.iloc[0].dropna().astype(str)[1:]
                ports = port_row.tolist()
                if not ports:
                    print(f"No PORTs found in '{sheet_name}' sheet")
                    return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), [f"No PORTs found in '{sheet_name}' sheet."]
                pol_row_idx = None
                for idx, row in df_raw.iterrows():
                    if str(row[0]).strip().upper() == "POL":
                        pol_row_idx = idx
                        break
                if pol_row_idx is None:
                    print(f"Could not detect POL row in '{sheet_name}' sheet")
                    return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), [f"Could not detect POL row in '{sheet_name}' sheet."]
                data_row_idx = pol_row_idx + 1
                data_row = df_raw.iloc[data_row_idx]
                pol = standardize_pol(str(data_row[0]).strip())
                if pol != "Nhava Sheva":
                    pol = "Nhava Sheva"
                records = []
                for port_idx, port in enumerate(ports):
                    col_20_idx = 1 + (port_idx * 2)
                    col_40_idx = col_20_idx + 1
                    rate_20 = data_row[col_20_idx] if col_20_idx < len(data_row) and pd.notna(data_row[col_20_idx]) else None
                    rate_40 = data_row[col_40_idx] if col_40_idx < len(data_row) and pd.notna(data_row[col_40_idx]) else None
                    rate_20_clean = clean_numeric_series(pd.Series([rate_20]))[0]
                    rate_40_clean = clean_numeric_series(pd.Series([rate_40]))[0]
                    if pd.isna(rate_20_clean) and pd.isna(rate_40_clean):
                        continue
                    records.append({
                        'POL': pol,
                        'PORT': port.title(),
                        '20': rate_20_clean,
                        '40HC': rate_40_clean,
                        'Remarks': ''
                    })
                for idx in range(data_row_idx + 1, len(df_raw)):
                    row = df_raw.iloc[idx]
                    for cell in row:
                        if pd.notna(cell) and str(cell).strip():
                            remarks.append(str(cell).strip())
                df_final = pd.DataFrame(records)
                if df_final.empty:
                    print(f"No valid records parsed from sheet '{sheet_name}'")
                    return df_final, remarks
                df_final['40STD'] = np.nan
                df_final = df_final[['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']]
                print(f"Parsed {len(df_final)} records from sheet '{sheet_name}', collected {len(remarks)} remarks")
                return df_final, remarks
            except Exception as e:
                print(f"Error in process_gulf_sheet for sheet '{sheet_name}': {str(e)}")
                return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), remarks + [f"Error in '{sheet_name}': {str(e)}"]

        def process_latam_sheet(sheet_name):
            remarks = []
            try:
                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None)
                records = []
                seen_lines = set()
                def is_record_row(row_vals):
                    return (
                        len(row_vals) >= 5 and
                        row_vals[0] and "/" in str(row_vals[0]) and
                        row_vals[2] and (
                            isinstance(row_vals[3], (int, float)) or
                            (isinstance(row_vals[3], str) and row_vals[3].replace(',', '').replace('.', '').isdigit()) or
                            row_vals[3] == 'NA'
                        ) and (
                            isinstance(row_vals[4], (int, float)) or
                            (isinstance(row_vals[4], str) and row_vals[4].replace(',', '').replace('.', '').isdigit()) or
                            row_vals[4] == 'NA'
                        )
                    )
                for idx, row in df_raw.iterrows():
                    row_vals = row.tolist()
                    full_line = " ".join(str(cell).strip() for cell in row_vals if pd.notna(cell) and str(cell).strip())
                    if is_record_row(row_vals):
                        raw_pol = str(row_vals[0]).strip()
                        port = str(row_vals[2]).strip()
                        rate_20 = clean_numeric_series(pd.Series([row_vals[3]]))[0]
                        rate_40 = clean_numeric_series(pd.Series([row_vals[4]]))[0]
                        if pd.isna(rate_20) and pd.isna(rate_40):
                            continue
                        for pol_code in raw_pol.split("/"):
                            pol = standardize_pol(pol_code.strip())
                            if pol not in ["Nhava Sheva", "Mundra", "Hazira"]:
                                continue
                            records.append({
                                'POL': pol,
                                'PORT': port,
                                '20': rate_20,
                                '40HC': rate_40,
                                '40STD': None,
                                'Remarks': ''
                            })
                    elif full_line and full_line not in seen_lines:
                        remarks.append(full_line)
                        seen_lines.add(full_line)
                df_final = pd.DataFrame(records)
                if df_final.empty:
                    print(f"No valid records parsed from sheet '{sheet_name}'")
                    return df_final, remarks
                pol_order = {'Nhava Sheva': 1, 'Mundra': 2, 'Hazira': 3}
                df_final['POL_ORDER'] = df_final['POL'].map(pol_order)
                df_final = df_final.sort_values(by='POL_ORDER').drop(columns='POL_ORDER')
                df_final = df_final[['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']]
                print(f"Parsed {len(df_final)} records from sheet '{sheet_name}', collected {len(remarks)} remarks")
                return df_final, remarks
            except Exception as e:
                print(f"Error in process_latam_sheet for sheet '{sheet_name}': {str(e)}")
                return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), remarks + [f"Error in '{sheet_name}': {str(e)}"]

        def process_canada_sheet(sheet_name):
            remarks = []
            try:
                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None, dtype=str)
                header_row_idx = None
                for idx, row in df_raw.iterrows():
                    if "DV20" in " ".join(row.dropna().astype(str).str.upper()) and "HC40" in " ".join(row.dropna().astype(str).str.upper()):
                        header_row_idx = idx
                        break
                if header_row_idx is None:
                    print(f"Header row not found in '{sheet_name}' sheet")
                    return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), [f"Header row not found in '{sheet_name}' sheet."]
                header_row = df_raw.iloc[header_row_idx]
                pol_row = df_raw.iloc[header_row_idx - 1]
                port_start_row = header_row_idx + 1
                all_pols = [standardize_pol(pol) for pol in pol_row if pd.notna(pol) and pol.strip() and standardize_pol(pol) != 'Wrg']
                rate_col_pairs = [(i, i + 1) for i in range(1, len(header_row)-1, 2)
                                if header_row[i].upper() == "DV20" and header_row[i + 1].upper() == "HC40"]
                ports = []
                data_end_row = None
                for idx in range(port_start_row, len(df_raw)):
                    port = str(df_raw.iloc[idx, 0]).strip()
                    if not port or any(word in port.upper() for word in ["INCLUSIVE", "SUBJECT", "NBF", "ZIM", "CN RAIL", "NOTE"]):
                        data_end_row = idx
                        break
                    ports.append(port)
                if data_end_row is None:
                    data_end_row = len(df_raw)
                for idx in range(0, port_start_row):
                    line = " ".join(str(cell).strip() for cell in df_raw.iloc[idx] if pd.notna(cell) and str(cell).strip())
                    if line:
                        remarks.append(line)
                for idx in range(data_end_row, len(df_raw)):
                    line = " ".join(str(cell).strip() for cell in df_raw.iloc[idx] if pd.notna(cell) and str(cell).strip())
                    if line:
                        remarks.append(line)
                records = []
                for pol_idx, pol in enumerate(all_pols):
                    if pol_idx >= len(rate_col_pairs):
                        continue
                    col_20_idx, col_40_idx = rate_col_pairs[pol_idx]
                    for i, port in enumerate(ports):
                        row_idx = port_start_row + i
                        rate_20_raw = df_raw.iloc[row_idx, col_20_idx] if col_20_idx < len(df_raw.columns) else None
                        rate_40_raw = df_raw.iloc[row_idx, col_40_idx] if col_40_idx < len(df_raw.columns) else None
                        rate_20 = clean_numeric_series(pd.Series([rate_20_raw]))[0] if rate_20_raw else None
                        rate_40 = clean_numeric_series(pd.Series([rate_40_raw]))[0] if rate_40_raw else None
                        if pd.notna(rate_20) or pd.notna(rate_40):
                            records.append({
                                'POL': pol,
                                'PORT': port,
                                '20': rate_20,
                                '40HC': rate_40,
                                '40STD': None,
                                'Remarks': ''
                            })
                df_final = pd.DataFrame(records)
                if df_final.empty:
                    print(f"No valid records parsed from sheet '{sheet_name}'")
                    return df_final, remarks
                df_final = df_final[['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']]
                print(f"Parsed {len(df_final)} records from sheet '{sheet_name}', collected {len(remarks)} remarks")
                return df_final, remarks
            except Exception as e:
                print(f"Error in process_canada_sheet for sheet '{sheet_name}': {str(e)}")
                return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), remarks + [f"Error in '{sheet_name}': {str(e)}"]

        def process_africa_sheet(sheet_name):
            remarks = []
            try:
                df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None)
                table_start_row = None
                for idx, row in df_raw.iterrows():
                    row_str = ' '.join(row.dropna().astype(str).str.upper())
                    if "MUNDRA" in row_str or "NHAVA SHEVA" in row_str or "PIPAVAV" in row_str:
                        table_start_row = idx
                        break
                if table_start_row is None:
                    print(f"Could not detect table start in '{sheet_name}' sheet")
                    return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), [f"Could not detect table start in '{sheet_name}' sheet."]
                for row_idx in range(table_start_row):
                    row = df_raw.iloc[row_idx]
                    for cell in row:
                        if pd.notna(cell) and str(cell).strip():
                            remarks.append(str(cell).strip())
                known_ports = ["APAPA", "TINCAN", "TEMA", "ABIDJAN", "COTONOU", "CONTONU", "LOME", "LEKKI", "ONNE", "MOMBASA", "DAR ES SALAAM"]
                port_row_idx = None
                for idx in range(table_start_row + 1, len(df_raw)):
                    row = df_raw.iloc[idx]
                    row_str = ' '.join(row.dropna().astype(str).str.upper())
                    if any(port in row_str for port in known_ports):
                        port_row_idx = idx
                        break
                if port_row_idx is None:
                    print(f"Could not detect PORT row in '{sheet_name}' sheet")
                    return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), [f"Could not detect PORT row in '{sheet_name}' sheet."]
                for row_idx in range(table_start_row + 1, port_row_idx):
                    row = df_raw.iloc[row_idx]
                    for cell in row:
                        if pd.notna(cell) and str(cell).strip():
                            remarks.append(str(cell).strip())
                pol_row = df_raw.iloc[table_start_row].dropna().astype(str).str.upper()
                port_row = df_raw.iloc[port_row_idx].dropna().astype(str)
                header_row = df_raw.iloc[port_row_idx + 1].dropna().astype(str)
                pols = [standardize_pol(p.strip()) for p in pol_row.iloc[0].split('/')] if not pol_row.empty else []
                ports = port_row.tolist()
                headers = header_row.tolist()
                if len(ports) == 0 or len(headers) == 0:
                    print(f"No PORTs or headers in '{sheet_name}' sheet")
                    return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), [f"No PORTs or headers in '{sheet_name}' sheet."]
                if "west africa" in sheet_name.lower().strip():
                    if len(headers) % 2 != 0 or len(headers) // 2 != len(ports):
                        print(f"Inconsistent PORT or header structure in '{sheet_name}' sheet")
                        return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), [f"Inconsistent PORT or header structure in '{sheet_name}' sheet."]
                else:
                    if len(ports) * 2 != len(headers):
                        print(f"Inconsistent PORT or header structure in '{sheet_name}' sheet")
                        return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), [f"Inconsistent PORT or header structure in '{sheet_name}' sheet."]
                data_start_row = port_row_idx + 2
                while data_start_row < len(df_raw):
                    data_row = df_raw.iloc[data_start_row]
                    if any(pd.to_numeric(str(cell), errors='coerce') == cell for cell in data_row.iloc[:len(headers)] if pd.notna(cell)):
                        break
                    for cell in data_row:
                        if pd.notna(cell) and str(cell).strip():
                            remarks.append(str(cell).strip())
                    data_start_row += 1
                if data_start_row >= len(df_raw):
                    print(f"No data rows found in '{sheet_name}' sheet")
                    return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), remarks
                data_rows = df_raw.iloc[data_start_row:data_start_row + 1]
                if data_rows.empty or len(data_rows.columns) < len(headers):
                    print(f"No valid data rows or insufficient columns in '{sheet_name}' sheet")
                    return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), remarks
                data_frames = []
                for port_idx, port in enumerate(ports):
                    dv20_value = data_rows.iloc[0, 2 * port_idx] if 2 * port_idx < len(data_rows.columns) else None
                    hc40_value = data_rows.iloc[0, 2 * port_idx + 1] if 2 * port_idx + 1 < len(data_rows.columns) else None
                    dv20_clean = clean_numeric_series(pd.Series([dv20_value]))[0]
                    hc40_clean = clean_numeric_series(pd.Series([hc40_value]))[0]
                    if pd.isna(dv20_clean) and pd.isna(hc40_clean):
                        continue
                    for pol in pols:
                        df_pol_port = pd.DataFrame({
                            'POL': [pol],
                            'PORT': [port.title()],
                            '20': [dv20_clean],
                            '40HC': [hc40_clean],
                            '40STD': [np.nan],
                            'Remarks': ['']
                        })
                        data_frames.append(df_pol_port)
                parsed_df = pd.concat(data_frames, ignore_index=True) if data_frames else pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks'])
                parsed_df['POL'] = parsed_df['POL'].apply(standardize_pol)
                parsed_df = parsed_df[parsed_df['PORT'].str.strip() != '']
                print(f"Parsed {len(parsed_df)} rows from sheet '{sheet_name}', collected {len(remarks)} remarks")
                return parsed_df, remarks
            except Exception as e:
                print(f"Error in process_africa_sheet for sheet '{sheet_name}': {str(e)}")
                return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), remarks + [f"Error in '{sheet_name}': {str(e)}"]

        for sheet_name in sheets:
            sheet_name_lower = sheet_name.lower().strip()
            if 'turkey' in sheet_name_lower and 'med' in sheet_name_lower:
                df, remarks = process_turkey_med_sheet(sheet_name)
            elif 'australia' in sheet_name_lower or 'aus' in sheet_name_lower:
                df, remarks = process_australia_sheet(sheet_name)
            elif 'ecsa' in sheet_name_lower:
                df, remarks = process_ecsa_sheet(sheet_name)
            elif 'gulf' in sheet_name_lower:
                df, remarks = process_gulf_sheet(sheet_name)
            elif 'latam' in sheet_name_lower:
                df, remarks = process_latam_sheet(sheet_name)
            elif 'canada' in sheet_name_lower:
                df, remarks = process_canada_sheet(sheet_name)
            elif 'africa' in sheet_name_lower:
                df, remarks = process_africa_sheet(sheet_name)
            elif 'usa' in sheet_name_lower:
                df, remarks = process_usa_sheet(sheet_name, file)
            else:
                df, remarks = process_generic_sheet(sheet_name, 'Nhava Sheva')

            # Append parsed data
            if not df.empty:
                data_frames.append(df)

            # Add remarks as usual
            all_remarks.extend([f"{sheet_name}: {remark}" for remark in remarks if remark and len(remark.split()) > 2])

            # ✅ Dynamically extract additional info from 4 specific sheets
            if sheet_name.strip().lower() in ['nhava sheva - far east', 'hazira - far east', 'east africa', 'west africa', 'usa']:
                try:
                    wb = load_workbook(file, data_only=True)
                    ws = wb[sheet_name]
                    seen_lines = set()

                    for row in ws.iter_rows():
                        cell_values = [cell.value for cell in row if cell.value is not None]
                        if not cell_values:
                            continue

                        # Check if row has any numeric-looking values (rates)
                        has_numeric = any(
                            isinstance(val, (int, float)) or
                            (isinstance(val, str) and any(char.isdigit() for char in val) and val.replace(',', '').replace('.', '', 1).isdigit())
                            for val in cell_values
                        )

                        # Skip rows that contain numeric values (likely record rows)
                        if has_numeric:
                            continue

                        # Combine and clean the row text
                        text_line = " ".join(str(val).strip() for val in cell_values if isinstance(val, str) and val.strip())
                        if text_line and text_line not in seen_lines and len(text_line.split()) >= 3:
                            all_remarks.append(f"{sheet_name}: {text_line}")
                            seen_lines.add(text_line)

                except Exception as e:
                    print(f"Failed to extract additional info from '{sheet_name}': {str(e)}")


        if not data_frames:
            print("No data frames parsed from any sheet")
            return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), all_remarks

        final_df = pd.concat(data_frames, ignore_index=True)
        final_df['POL'] = final_df['POL'].apply(standardize_pol)
        final_df = final_df[final_df['PORT'].notna() & (final_df['PORT'].str.strip() != '')]
        final_df = final_df.sort_values(by=['POL', 'PORT'])
        final_df = format_rates_with_dollar(final_df)
        all_remarks = list(dict.fromkeys(all_remarks))
        final_df = split_ports(final_df, port_column='PORT')

        print(f"Total records parsed: {len(final_df)}")
        print(f"Total remarks collected: {len(all_remarks)}")
        return final_df, all_remarks

    except Exception as e:
        print(f"Error parsing ZIM MRG file: {str(e)}")
        print(f"Full traceback: {traceback.format_exc()}")
        return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), [f"Error parsing ZIM MRG file: {str(e)}"]

def process_usa_sheet(sheet_name, file):
    remarks = []
    try:
        df_raw = pd.read_excel(file, sheet_name=sheet_name, header=None)
        records = []
        seen_lines = set()
        pols = ['Nhava Sheva', 'Mundra', 'Pipavav']
        port_row_keywords = ['NEW', 'NORFOLK', 'MIAMI', 'BOSTON', 'HOUSTON', 'BALTIMORE', 'LOS', 'TAMPA', 'MOBILE', 'ORLEANS']

        for idx, row in df_raw.iterrows():
            cells = row.dropna().astype(str).str.strip().tolist()
            if not cells:
                continue

            first_cell_upper = cells[0].upper() if cells else ''

            # Identify valid port rows (first cell looks like a US port)
            is_port_row = any(first_cell_upper.startswith(k) for k in port_row_keywords)
            has_rate_like_numbers = sum(1 for c in cells if any(char.isdigit() for char in c)) >= 2

            if is_port_row:
                port = cells[0].split(',')[0].strip().title()
                last_3 = cells[-3:] if len(cells) >= 3 else [None, None, None]
                rates = [clean_numeric_series(pd.Series([v]))[0] for v in last_3]

                for pol in pols:
                    records.append({
                        'POL': pol,
                        'PORT': port,
                        '20': rates[0] if len(rates) > 0 else None,
                        '40STD': rates[1] if len(rates) > 1 else None,
                        '40HC': rates[2] if len(rates) > 2 else None,
                        'Remarks': ''
                    })
            else:
                # Pure remark line (not a port name, not numeric-heavy)
                line = ' '.join(cells)
                if len(line.split()) >= 3 and line not in seen_lines:
                    remarks.append(line)
                    seen_lines.add(line)

        df_final = pd.DataFrame(records)
        df_final = df_final[['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']]
        print(f"Parsed {len(df_final)} USA records, collected {len(remarks)} remarks")
        df_final = split_ports(df_final, port_column='PORT')
        return df_final, remarks

    except Exception as e:
        print(f"Error in 'USA': {str(e)}")
        return pd.DataFrame(columns=['POL', 'PORT', '20', '40STD', '40HC', 'Remarks']), [f"Error in 'USA': {str(e)}"]
    
def parse_custom_vendor(file, month_year, selected_columns):
    try:
        # Read Excel file (first sheet, assume headers)
        df_raw = pd.read_excel(file, sheet_name=0, header=0)
        df_raw.columns = df_raw.columns.astype(str).str.strip()

        # Validate required columns
        if 'POL' not in selected_columns or 'PORT' not in selected_columns:
            return pd.DataFrame(), ["Error: 'POL' and 'PORT' columns are required."]

        # Map Excel columns to selected columns (case-insensitive, partial match)
        col_map = {}
        for sel_col in selected_columns:
            for raw_col in df_raw.columns:
                if sel_col.lower() in raw_col.lower():
                    col_map[raw_col] = sel_col
                    break

        # Create DataFrame with mapped columns
        df = pd.DataFrame()
        if col_map:
            df = df_raw[list(col_map.keys())].rename(columns=col_map)
        else:
            df = df_raw.copy()

        # Ensure all selected columns exist
        for col in selected_columns:
            if col not in df.columns:
                df[col] = np.nan

        # Reorder columns
        df = df[selected_columns]

        # Clean and standardize data with currency formatting
        numeric_cols = ['20', '40STD', '40HC']
        for col in numeric_cols:
            if col in df.columns:
                def format_currency(value):
                    if pd.isna(value):
                        return value
                    value_str = str(value).strip().upper()
                    numeric_value = clean_numeric(pd.Series([value]))[0]
                    if pd.isna(numeric_value):
                        return None
                    if 'EUR' in value_str or '€' in value_str:
                        return f"€{int(numeric_value)}"  # Convert to integer
                    else:
                        return f"${int(numeric_value)}"  # Convert to integer
                df[col] = df[col].apply(format_currency)

        text_cols = ['POL', 'PORT', 'REMARKS', 'ROUTING', 'TRANSIT TIME', 'SERVICE', 'VALIDITY']
        for col in text_cols:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip().replace('nan', '')
                if col == 'POL':
                    df['POL'] = df['POL'].apply(standardize_pol)

        if 'EXPIRY DATE' in df.columns:
            df['EXPIRY DATE'] = pd.to_datetime(df['EXPIRY DATE'], errors='coerce').dt.strftime('%Y-%m-%d')

        # Filter out invalid rows
        df = df[df['PORT'].notna() & (df['PORT'].str.strip() != '')]
        df = df[df['POL'].notna() & (df['POL'].str.strip() != '')]

        # Split ports in the PORT column
        df = split_ports(df, port_column='PORT')

        # Extract remarks from non-data sections
        remarks = []
        try:
            df_header = pd.read_excel(file, sheet_name=0, header=None)
            for idx, row in df_header.iterrows():
                for cell in row:
                    if pd.notna(cell) and str(cell).strip():
                        cell_text = str(cell).strip()
                        if len(cell_text.split()) > 2 and not any(cell_text.lower() in str(val).lower() for val in df['PORT']):
                            remarks.append(cell_text)
        except Exception as e:
            remarks.append(f"Error extracting remarks: {str(e)}")

        return df, remarks
    except Exception as e:
        print(f"Error parsing custom vendor file: {str(e)}")
        return pd.DataFrame(), [f"Error parsing custom vendor file: {str(e)}"]
    
# For demonstration, a subset of the mapping is included as a dictionary
port_mapping = {
    'abidjan': 'Abidjan',
    'abu dhabi': 'Abu Dhabi',
    'abu dhabi, abu zabi, united arab emirates': 'Abu Dhabi',
    'acajutla': 'Acajutla, El Salvador',
    'acajutla, el salvador': 'Acajutla, El Salvador',
    'adelaide': 'Adelaide',
    'aden': 'Aden',
    'agadir': 'Agadir',
    'ajman': 'Ajman',
    'ajman, ras al khaymah, united arab emirates': 'Ajman',
    'al sokhna': 'Sokhna',
    'alexandria': 'Alexandria',
    'alexandria (el dekheila)': 'Alexandria',
    'algiers': 'Algiers',
    'aliaga': 'Aliaga',
    'altamira': 'Altamira, MX',
    'altamira, mexico': 'Altamira, MX',
    'annaba': 'Annaba',
    'anqing, anqing, anhui, china': 'Anqing',
    'antwerp': 'Antwerp',
    'apapa': 'Apapa',
    'aqaba': 'Aqaba',
    'arica': 'Arica',
    'arica, chile': 'Arica',
    'auckland': 'Auckland',
    'bahrain': 'Bahrain',
    'bahrain, bahrain': 'Bahrain',
    'balboa,panama': 'Balboa',
    'balikpapan': 'Balikpapan',
    'baltimore': 'Baltimore',
    'bangkok': 'Bangkok',
    'bangkok (bmt)': 'Bangkok',
    'bangkok (pat)': 'Bangkok PAT',
    'bangkok pat': 'Bangkok PAT',
    'bangkok, thailand': 'Bangkok',
    'bangkok,pat': 'Bangkok PAT',
    'banjul': 'Banjul',
    'baranquilla, colombia': 'Baranquilla',
    'barcelona': 'Barcelona',
    'batam': 'Batam',
    'beicun': 'Beicun',
    'beihai, beihai, guangxi, china': 'Beihai',
    'beijao': 'Beijao',
    'beijiao, shunde, guangdong, china': 'Beijao',
    'beira': 'Beira',
    'beirut': 'Beirut',
    'bejaia': 'Bejaia',
    'belawan': 'Belawan',
    'belawan, sumatera utara, indonesia': 'Belawan',
    'bell bay': 'Bell bay',
    'benghazi': 'Benghazi',
    'berbera': 'Berbera',
    'bintulu': 'Bintulu',
    'bissau': 'Bissau',
    'bluff': 'Bluff',
    'boston': 'Boston',
    'bremerhaven': 'Bremerhaven',
    'bridgetown': 'Bridgetown',
    'brisbane': 'Brisbane',
    'buenaventura': 'Buenaventura',
    'buenaventura,colombia': 'Buenaventura',
    'buenos aires': 'Buenos Aires',
    'busan': 'Busan',
    'busan, south korea': 'Busan',
    'caacupemi asuncion': 'Asuncion',
    'caacupemi pilar': 'Pilar',
    'cagayan de oro': 'Cagayan De Oro',
    'cai mep': 'Cai Mep',
    'callao': 'Callao',
    'callao, peru': 'Callao',
    'cape town': 'Cape Town',
    'capetown': 'Cape Town',
    'cartagena': 'Cartagena',
    'cartagena, colombia': 'Cartagena',
    'casablanca': 'Casablanca',
    'cat lai, hcmc': 'Cat Lai',
    'catlai': 'Cat Lai',
    'caucedo': 'Caucedo',
    'caucedo, dominican republic': 'Caucedo',
    'cebu': 'Cebu',
    'cebu city, cebu, philippines': 'Cebu City',
    'chancay, peru': 'Chancay',
    'changsha': 'Changsha',
    'changsha, changsha, hunan, china': 'Changsha',
    'changzhou': 'Changzhou',
    'changzhou, changzhou, jiangsu, china': 'Changzhou',
    'charleston': 'Charleston',
    'chattogram': 'Chattogram',
    'chittagong': 'Chittagong',
    'chittagong *(incl dthc)': 'Chittagong',
    'chiwan, shekou': 'Chiwan',
    'chongqing': 'Chongqing',
    'chongqing, china': 'Chongqing',
    'coega': 'Coega',
    'colombo': 'Colombo',
    'colombo *(incl. dthc': 'Colombo',
    'colombo, western, sri lanka': 'Colombo',
    'colon free zone, door': 'Colon Free Zone',
    'conakry': 'Conakry',
    'constanta': 'Constanta',
    'corinto': 'Corinto',
    'corinto, nicaragua': 'Corinto',
    'coronel': 'Coronel',
    'cotonou': 'Cotonou',
    'cristobal': 'Cristobal',
    'dakar': 'Dakar',
    'dalian': 'Dalian',
    'dalian, dalian, liaoning, china': 'Dalian',
    'dammam': 'Dammam',
    'dammam bonded re-export zone, ash sharqiyah, saudi arabia': 'Dammam Bonded Zone',
    'dammam, ash sharqiyah, saudi arabia': 'Dammam',
    'danang': 'Danang',
    'dar es salaam': 'Dar es salaam',
    'dar-es-salaam': 'Dar es salaam',
    'davao': 'Davao',
    'diego suarez': 'Diego suarez',
    'djibouti': 'Djibouti',
    'djjib': 'Djibouti',
    'dongguan, dongguan, guangdong, china': 'Dongguan',
    'dou men, zuhai province': 'Dou men, Zuhai Province',
    'douala': 'Douala',
    'dubai, dubai, united arab emirates': 'Jebel Ali',
    'durban': 'Durban',
    'dzaae': 'Annaba',
    'dzalg': 'Alger',
    'dzorn': 'Oran',
    'dzski': 'Skikda',
    'egaly': 'Alexandria',
    'egpsd': 'Port Said',
    'ensenada': 'Ensenada',
    'ensenada,mexico': 'Ensenada',
    'esbcn': 'Barcelona',
    'esvlc': 'Valencia',
    'fangcheng, fangchenggang, guangxi, china': 'Fangcheng',
    'felixstowe': 'Felixstowe',
    'fos sur mer': 'Fos sur mer',
    'foshan jiujiang': 'Foshan jiujiang',
    'foshan, foshan, guangdong, china': 'Foshan',
    'freeport': 'Freeport',
    'freetown': 'Freetown',
    'fremantle': 'Fremantle',
    'frfos': 'Fos Sur Mer',
    'fukuyama': 'Fukuyama',
    'fuqing, fuzhou, fujian, china': 'Fuqing',
    'fuzhou, fuzhou, fujian, china': 'Fuzhou',
    'gao ming': 'Gao ming',
    'gao yao': 'Gao yao',
    'gaolan, zhuhai, zhuhai, guangdong, china': 'Gaolan, Zhuhai',
    'gaolan, zuhai province': 'Gaolan, Zhuhai',
    'gaoming, foshan, guangdong, china': 'Gaoming, Foshan, Guangdong',
    'gaosha': 'Gaosha',
    'gemlik': 'Gemlik',
    'genoa': 'Genoa',
    'georgetown': 'Georgetown',
    'georgetown, guyana': 'Georgetown',
    'grpir': 'Piraeus',
    'grskg': 'Thessaloniki',
    'guangzhou, guangzhou, guangdong, china': 'Guangzhou, Guangzhou, Guangdong',
    'guatemala city': 'Guatemala city',
    'guayaquil': 'Guayaquil',
    'guayaquil, equador': 'Guayaquil',
    'hai phong': 'Haiphong',
    'hai phong, vietnam': 'Haiphong',
    'haikou': 'Haikou',
    'haikou, haikou, hainan, china': 'Haikou, Haikou, Hainan',
    'haiphong': 'Haiphong',
    'hakata': 'Hakata',
    'halifax': 'Halifax',
    'hamad': 'Hamad',
    'hamad, qatar': 'Hamad',
    'hamburg': 'Hamburg',
    'haungpu': 'Haungpu',
    'hefei': 'Hefei',
    'hefei, hefei, anhui, china': 'Hefei',
    'hiroshima': 'Hiroshima',
    'ho chi minh': 'Ho Chi Minh City',
    'ho chi minh, vietnam': 'Ho Chi Minh City',
    'hochiminh vict': 'Ho Chi Minh City, VICT',
    'hong kong': 'Hong Kong',
    'hong kong, hong kong, china': 'Hong Kong',
    'hongkong': 'Hong Kong',
    'hongwan, zuhai province': 'Hongwan, zuhai province',
    'houston': 'Houston',
    'huadu': 'Huadu',
    'huadu, guangzhou, guangdong, china': 'Huadu, Guangzhou, Guangdong',
    'huangpu': 'Huangpu',
    'huangpu, guangzhou, guangdong, china': 'Huangpu, Guangzhou, Guangdong',
    'humen': 'Humen',
    'humen, dongguan, guangdong, china': 'Humen, Dongguan, Guangdong',
    'icd dhaka': 'Icd dhaka',
    'icd phuoc long 1': 'Icd phuoc long 1',
    'icd phuoc long 3': 'Icd phuoc long 3',
    'imbitumba': 'Imbitumba',
    'incheon': 'Incheon',
    'inchon': 'Incheon',
    'inchon, south korea': 'Incheon',
    'iquique': 'Iquique',
    'iquique, chile': 'Iquique',
    'iskenderun': 'Iskenderun',
    'istanbul': 'Istanbul',
    'istanbul (ambarli-kumport)': 'Istanbul',
    'istanbul (haydarpasa)': 'Istanbul',
    'itajai': 'Itajai',
    'itapoa': 'Itapoa',
    'itgoa': 'Genoa',
    'itsal': 'Salerno',
    'itspe': 'La spezia',
    'izmir': 'Izmir',
    'izmit': 'Izmit',
    'izmit (evyap)': 'Izmit',
    'jacksonville': 'Jacksonville',
    'jakarta': 'Jakarta',
    'jakarta, indonesia': 'Jakarta',
    'jebel ali': 'Jebel Ali',
    'jebel ali, dubai, united arab emirates': 'Jebel Ali',
    'jeddah': 'Jeddah',
    'jiangmen new port': 'Jiangmen',
    'jiangmen, jiangmen, guangdong, china': 'Jiangmen',
    'jiangyin, wuxi, jiangsu, china': 'Jiangyin, Wuxi, Jiangsu',
    'jiangyin,fujian': 'Jiangyin,Fujian',
    'jiaoxin': 'Jiaoxin',
    'jiaoxin, guangzhou, guangdong, china': 'Jiaoxin',
    'jiaxing, jiaxing, zhejiang, china': 'Jiaxing',
    'jing zhou': 'Jingzhou',
    'jingzhou, jingzhou, hubei, china': 'Jingzhou',
    'jinzhou': 'Jingzhou',
    'jiujiang': 'Jiujiang',
    'jiujiang, jiujiang, jiangxi, china': 'Jiujiang',
    'joaqj': 'Aqaba',
    'jubail': 'Jubail',
    'jubail, ash sharqiyah, saudi arabia': 'Jubail',
    'kaiping, jiangmen, guangdong, china': 'Kaiping, Jiangmen, Guangdong',
    'kaohsiung': 'Kaohsiung',
    'kaohsiung, taiwan': 'Kaohsiung',
    'kaouhsiung': 'Kaohsiung',
    'kawasaki': 'Kawasaki',
    'keelung': 'Keelung',
    'keelung, taiwan': 'Keelung',
    'khoms': 'Khoms',
    'king abdullah port - kap': 'King abdullah port',
    'kingston': 'Kingston',
    'kingston, jamaica': 'Kingston',
    'kobe': 'Kobe',
    'kobe, hyogo-ken, japan': 'Kobe',
    'kota kinabalu': 'Kota kinabalu',
    'kribi': 'Kribi',
    'kuching': 'Kuching',
    'kwangyang': 'Kwangyang',
    'la guaira - euro': 'La Guaira - EURO',
    'la spezia': 'La spezia',
    'laem cha bang': 'Laem cha bang',
    'laem chabang': 'Laem cha bang',
    'laem chabang, chon buri, thailand': 'Laem cha bang',
    'lanshi, foshan, guangdong, china': 'All Separate',
    'lat kra bang': 'Lat Krabang',
    'lat krabang': 'Lat Krabang',
    'lat krabang, bangkok, thailand': 'Lat Krabang',
    'latkra bang': 'Lat Krabang',
    'lautoka': 'Lautoka',
    'lazaro cardenas': 'Lazaro cardenas',
    'lazaro cardenas, mexico': 'Lazaro cardenas',
    'lbbey': 'Beirut',
    'lekki': 'Lekki',
    'lian hua shan': 'Lian hua shan',
    'lianhuashan, guangzhou, guangdong, china': 'Lian hua shan',
    'lianyungang': 'Lianyungang',
    'lianyungang, lianyungang, jiangsu, china': 'Lianyungang',
    'libreville': 'Libreville',
    'lirquen chile': 'Lirquen',
    'liudu, yunfu, guangdong, china': 'Liudu',
    'lobito': 'Lobito',
    'lome': 'Lome',
    'long beach': 'Long beach',
    'longoni': 'Longoni',
    'luanda': 'Luanda',
    'luzhou, luzhou, sichuan, china': 'Luzhou',
    'lyben': 'Benghazi',
    'lymra': 'Misurata',
    'lytip': 'Tripoli',
    'lyttleton': 'Lyttleton',
    'macas': 'Casablanca',
    'majunga': 'Majunga',
    'male *(incl pad)': 'Male',
    'manaus': 'Manaus',
    'manila north': 'Manila north',
    'manila south': 'Manila south',
    'manila, metro manila, philippines': 'Manila',
    'manila,north': 'Manila north',
    'manzanillo': 'Manzanillo, MX',
    'manzanillo, panama': 'Manzanillo, PA',
    'manzanillo,mx': 'Manzanillo, MX',
    'maputo': 'Maputo',
    'matadi': 'Matadi',
    'matsuyama': 'Matsuyama',
    'mawei, fuzhou': 'Mawei',
    'mawei, fuzhou, fujian, china': 'Mawei',
    'melbourne': 'Melbourne',
    'mersin': 'Mersin',
    'mexico city': 'Mexico city',
    'mindelo': 'Mindelo',
    'miri': 'Miri',
    'misurata': 'Misurata',
    'mizushima': 'Mizushima',
    'mobile': 'Mobile',
    'mogadishu': 'Mogadishu',
    'moin': 'Moin',
    'moji': 'Moji',
    'mombasa': 'Mombasa',
    'monrovia': 'Monrovia',
    'montevideo': 'Montevideo',
    'montreal': 'Montreal',
    'moroni': 'Moroni',
    'mukalla': 'Mukalla',
    'nacala': 'Nacala',
    'nador': 'Nador',
    'nagoya': 'Nagoya',
    'nagoya, aichi, japan': 'Nagoya',
    'naha okinawa': 'Naha okinawa',
    'namibe': 'Namibe',
    'nanchang': 'Nanchang',
    'nanchang, nanchang, jiangxi, china': 'Nanchang',
    'nanjing': 'Nanjing',
    'nanjing, nanjing, jiangsu, china': 'Nanjing',
    'nansha': 'Nansha',
    'nansha new port': 'Nansha New Port',
    'nansha, guangzhou, guangdong, china': 'Nansha',
    'nantong': 'Nantong',
    'nantong, nantong, jiangsu, china': 'Nantong',
    'napier': 'Napier',
    'nassau': 'Nassau',
    'navegantes': 'Navegantes',
    'nelson': 'Nelson',
    'new orleans': 'New orleans',
    'new york': 'New york',
    'ningbo': 'Ningbo',
    'ningbo, ningbo, zhejiang, china': 'Ningbo',
    'norfolk': 'Norfolk',
    'nouadhibou': 'Nouadhibou',
    'nouakchott': 'Nouakchott',
    'noumea': 'Noumea',
    'oakland': 'Oakland',
    'onne': 'Onne',
    'oran': 'Oran',
    'osaka': 'Osaka',
    'osaka, japan': 'Osaka',
    'paita': 'Paita',
    'palembang': 'Palembang',
    'panama city, door': 'Panama City',
    'panjang': 'Panjang',
    'pantaco': 'Pantaco',
    'paramaribo': 'Paramaribo',
    'paramaribo, suriname': 'Paramaribo',
    'paranagua': 'Paranagua',
    'pasir gudang': 'Pasir gudang',
    'pasir gudang, johor, malaysia': 'Pasir gudang',
    'pasirgudang': 'Pasir gudang',
    'pecem': 'Pecem',
    'penang': 'Penang',
    'penang, pulau pinang, malaysia': 'Penang',
    'philadelphia': 'Philadelphia',
    'phnompenh': 'Phnom Penh',
    'pointe de galets': 'Pointe de galets',
    'pointe noire': 'Pointe noire',
    'pontianak': 'Pontianak',
    'port au prince': 'Port au prince',
    'port au prince,haiti': 'Port au prince',
    'port chalmers': 'Port chalmers',
    'port everglades': 'Port everglades',
    'port kelang': 'Port Klang',
    'port klang (north)': 'Port Klang North',
    'port klang (west)': 'Port Klang West',
    'port klang, selangor, malaysia': 'Port Klang',
    'port louis': 'Port louis',
    'port of spain': 'Port of spain',
    'port of spain, trindad & tobago': 'Port of Spain',
    'port said': 'Port said',
    'port sudan': 'Port sudan',
    'posorja,ecuador': 'Posorja',
    'poti': 'Poti',
    'praia': 'Praia',
    'psa, dongguan': 'Psa, dongguan',
    'puerto barrios': 'Puerto barrios',
    'puerto cabello - euro': 'Puerto Cabello - EURO',
    'puerto caldera': 'Puerto caldera',
    'puerto caldera, costa rica': 'Puerto caldera',
    'puerto cortes': 'Puerto cortes',
    'puerto cortes,honduras': 'Puerto Cortes',
    'puerto limon, costa rica': 'Puerto Limon',
    'puerto quetzal': 'Puertal Quetzal',
    'puerto quetzal, guatemala': 'Puertal Quetzal',
    'pusan': 'Pusan',
    'qingdao': 'Qingdao',
    'qingdao, qingdao, shandong, china': 'Qingdao',
    'qingyuan': 'Qingyuan',
    'qinzhou': 'Qinzhou',
    'qinzhou, qinzhou, guangxi, china': 'Qinzhou',
    'qui nhon': 'Qui nhon',
    'quingdao': 'Qingdao',
    'quy nhon': 'Qui nhon',
    'ras al khaimah': 'Ras al khaimah',
    'ras al khaimah, ras al khaymah, united arab emirates': 'Ras al khaimah',
    'rio de janeiro': 'Rio De Janeiro',
    'rio grande': 'Rio Grande',
    'rio haina': 'Rio Haina',
    'rio haina, dominican republic': 'Rio Haina',
    'riyadh': 'Riyadh',
    'riyadh, ar riyad, saudi arabia': 'Riyadh',
    'rodman': 'Rodman',
    'rongqi, shunde, guangdong, china': 'Rongqi, Shunde',
    'rosario': 'Rosario',
    'sajed': 'Jeddah',
    'salalah': 'Salalah',
    'salalah, zufar, oman': 'Salalah',
    'salvador (de bahia)': 'Salvador',
    'samut prakan (bangkok), samut prakan, thailand': 'Bangkok (Samut Prakan)',
    'san antonio': 'San antonio',
    'san antonio, chile': 'San Antonio',
    'san juan, puerto rico': 'San Juan',
    'san lorenzo, honduras': 'San Lorenzo',
    'sandakan': 'Sandakan',
    'shanghai': 'Shanghai',
    'san-pedro': 'San-pedro',
    'sanshan': 'Sanshan',
    'sanshan, foshan, guangdong, china': 'Sanshan',
    'sanshui new port': 'Sanshui New Port',
    'sanshui, foshan, guangdong, china': 'Sanshui',
    'santo tomas de castilla, guatemala': 'Santo tomas de castilla',
    'santos': 'Santos',
    'savannah': 'Savannah',
    'sdpzu': 'Port Sudan',
    'seattle': 'Seattle',
    'semarang': 'Semarang',
    'semarang, jawa tengah, indonesia': 'Semarang',
    'sendai, miyagi': 'Sendai',
    'shanghai, china': 'Shanghai',
    'shantou, shantou, guangdong, china': 'Shantou',
    'sharjah': 'Sharjah',
    'sharjah, sharjah, united arab emirates': 'Sharjah',
    'shatian': 'Shatian',
    'shatian, dongguan, guangdong, china': 'Shatian, dongguan',
    'shekhou': 'Shekou',
    'shekou': 'Shekou',
    'shekou, shenzhen, china': 'Shekou',
    'shekou, shenzhen, guangdong, china': 'Shekou',
    'shimizu': 'Shimizu',
    'shuaiba': 'Shuaiba',
    'shuaiba, kuwait': 'Shuaiba',
    'shunde leliu wharf': 'Shunde Leliu Wharf',
    'shunde new port': 'Shunde new port',
    'shunde, shunde, guangdong, china': 'Shunde',
    'shuwaikh': 'Shuwaikh',
    'shuwaikh, kuwait': 'Shuwaikh',
    'si hui (ma fang)': 'Si Hui',
    'sibu': 'Sibu',
    'sihanoukville': 'Sihanoukville',
    'sihanoukville, sihanoukville, cambodia': 'Sihanoukville',
    'singapore': 'Singapore',
    'singapore, singapore': 'Singapore',
    'skikda': 'Skikda',
    'sohar': 'Sohar',
    'sohar, masqat, oman': 'Sohar',
    'sokhna port': 'Sokhna port',
    'suape': 'Suape',
    'subic bay': 'Subic bay',
    'surabaya': 'Surabaya',
    'surabaya, jawa timur, indonesia': 'Surabaya',
    'suva': 'Suva',
    'sydney': 'Sydney',
    'syltk': 'Latakia',
    'taicang': 'Taicang',
    'taicang, suzhou, jiangsu, china': 'Taicang',
    'taichung': 'Taichung',
    'taichung, taiwan': 'Taichung',
    'taipei': 'Taipei',
    'takoradi': 'Takoradi',
    'tamatave': 'Tamatave',
    'tanga': 'Tanga',
    'taoyuan': 'Taoyuan',
    'tauranga': 'Tauranga',
    'tawao': 'Tawao',
    'tema': 'Tema',
    'tianjin, xingang': 'Tianjin, Xingang',
    'tincan': 'Tincan',
    'tnsfa': 'Sfax',
    'tnsus': 'Sousse',
    'tokuyama': 'Tokuyama',
    'tokyo': 'Tokyo',
    'tokyo, japan': 'Tokyo',
    'tongling, tongling, anhui, china': 'Tongling',
    'toronto': 'Toronto',
    'trege': 'Eregli',
    'trgeb': 'Gebze',
    'trgem': 'Gemlik',
    'trisk': 'Iskenderun',
    'trmer': 'Mersin',
    'trmrp': 'Marport',
    'tunis': 'Tunis',
    'ulsan': 'Ulsan',
    'umm al qaiwain': 'Umm Al Qawain',
    'umm al qawain, umm al qaywayn, united arab emirates': 'Umm Al Qawain',
    'umm qasr *(incl. dthc)': 'Umm Qasr',
    'umm qasr north port, iraq': 'Umm Qasr North',
    'valencia': 'Valencia',
    'valparaiso, chile': 'Valparaiso',
    'vancouver': 'Vancouver',
    'veracruz': 'Veracruz',
    'veracruz, mexico': 'Veracruz',
    'vila do conde': 'Vila do conde',
    'vitoria': 'Vitoria',
    'walvis bay': 'Walvis bay',
    'wellington': 'Wellington',
    'wenzhou': 'Wenzhou',
    'wenzhou, wenzhou, zhejiang, china': 'Wenzhou',
    'wu zhou': 'Wu zhou',
    'wuhan': 'Wuhan',
    'wuhan, wuhan, hubei, china': 'Wuhan',
    'wuhu': 'Wuhu',
    'wuhu, wuhu, anhui, china': 'Wuhu',
    'wuxi': 'Wuxi',
    'xiamen': 'Xiamen',
    'xiamen, xiamen, fujian, china': 'Xiamen',
    'xiaolan': 'Xiaolan',
    'xinfeng, liwan, guangzhou, guangdong, china': 'Xinfeng, liwan',
    'xingang': 'Xingang',
    'xingang, tianjin, china': 'Xingang, Tianjin',
    'xinhui': 'Xinhui',
    'xinhui, jiangmen, guangdong, china': 'Xinhui, Jiangmen',
    'yang zhou': 'Yang zhou',
    'yangon': 'Yangon',
    'yangon, yangon, myanmar': 'Yangon',
    'yangpu pt': 'Yangpu',
    'yangpu, danzhou, hainan, china': 'Yangpu',
    'yangzhou, yangzhou, jiangsu, china': 'Yangzhou',
    'yantai, yantai, shandong, china': 'Yantai',
    'yantian': 'Yantian',
    'yantian, shenzhen, china': 'Yantian',
    'yantian, shenzhen, guangdong, china': 'Yantian',
    'yibin, yibin, sichuan, china': 'Yibin',
    'yichang': 'Yichang',
    'yichang, yichang, hubei, china': 'Yichang',
    'yokkaichi': 'Yokkaichi',
    'yokkaichi, mie, japan': 'Yokkaichi',
    'yokohama': 'Yokohama',
    'yokohama, kanagawa-ken, japan': 'Yokohama',
    'yueyang': 'Yueyang',
    'yueyang, yueyang, hunan, china': 'Yueyang',
    'zanzibar': 'Zanzibar',
    'zarate': 'Zarate',
    'zhangjiagang': 'Zhangjiagang',
    'zhanjiang': 'Zhangjiagang',
    'zhanjiang, zhanjiang, guangdong, china': 'Zhangjiagang',
    'zhaoqing': 'Zhaoqing',
    'zhaoqing new port': 'Zhaoqing New Port',
    'zhaoqing, zhaoqing, guangdong, china': 'Zhaoqing',
    'zhapu': 'Zhapu',
    'zhapu, pinghu, jiaxing, zhejiang, china': 'Zhapu',
    'zhongshan': 'Zhongshan',
    'zhongshan, zhongshan, guangdong, china': 'Zhongshan',
    'zhuhai, zhuhai, guangdong, china': 'Zhuhai'
}

def normalize_text(text):
    """
    Normalize text for consistent comparison (lowercase, strip whitespace).
    """
    if pd.isna(text):
        return ''
    return ' '.join(str(text).lower().strip().split())

def map_port_name(port):
    """
    Map a port name to its standardized name using the port_mapping dictionary.
    If no match is found, return the original port name with proper capitalization.
    """
    if not port or pd.isna(port):
        return ''
    normalized_port = normalize_text(port)
    mapped_port = port_mapping.get(normalized_port, port)
    # Capitalize first letter, lowercase the rest
    if mapped_port:
        return mapped_port[0].upper() + mapped_port[1:].lower() if len(mapped_port) > 1 else mapped_port.upper()
    return port[0].upper() + port[1:].lower() if len(port) > 1 else port.upper()

def split_ports(df, port_column='PORT'):
    """
    Split rows in a DataFrame where the port_column contains '/' or ';' into separate records.
    Map each port name to its standardized name from the master sheet.
    Format all port names to have only the first letter capitalized, rest in lowercase.
    
    Args:
        df (pd.DataFrame): Input DataFrame
        port_column (str): Column name containing port names (e.g., 'PORT' or 'POD')
    
    Returns:
        pd.DataFrame: DataFrame with split and mapped port records
    """
    if port_column not in df.columns:
        return df
    
    # Initialize a list to hold the new rows
    new_rows = []
    
    for idx, row in df.iterrows():
        # Split ports by '/' or ';'
        ports = str(row[port_column]).split('/')
        ports = [port for sublist in [p.split(';') for p in ports] for port in sublist]
        ports = [port.strip() for port in ports if port.strip()]
        
        # If no ports after splitting, skip the row
        if not ports:
            continue
            
        # Create a new row for each port with mapped and formatted name
        for port in ports:
            # Map the port name using the master sheet mapping
            mapped_port = map_port_name(port)
            new_row = row.copy()
            new_row[port_column] = mapped_port
            new_rows.append(new_row)
    
    # Create a new DataFrame with the expanded rows
    if new_rows:
        new_df = pd.DataFrame(new_rows).reset_index(drop=True)
        return new_df
    return df

# === FIREBASE DATA FUNCTIONS ===
@st.cache_data
def load_from_firestore():
    # Original hardcoded vendors
    data = {
        "MSC": {}, "Wan Hai": {}, "Emirates": {}, "ONE MRG": {}, "HMM MRG": {},
        "OOCL": {}, "PIL MRG": {}, "ARKAS MRG": {}, "Interasia": {}, "Cosco-Gulf": {}, 
        "Cosco-WCSA & CB": {}, "Cosco-Africa": {}, "Turkon": {}, "Cosco-Fareast": {}, "ZIM MRG": {}, "MSC-EUR MED": {}
    }
    all_pods = set()
    all_pols = set()
    vendors_ref = db.collection('vendors')
    
    # Process hardcoded vendors as before
    for vendor in data.keys():
        docs = vendors_ref.document(vendor).collection('data').stream()
        for doc in docs:
            data[vendor][doc.id] = doc.to_dict()
            df_temp = pd.DataFrame(doc.to_dict().get("data", []))
            if 'PORT' in df_temp.columns:
                all_pods.update(df_temp['PORT'].dropna().astype(str).str.strip().unique())
            if 'POD' in df_temp.columns:  # Handle HMM MRG case before renaming
                all_pods.update(df_temp['POD'].dropna().astype(str).str.strip().unique())
            if 'POL' in df_temp.columns:
                df_temp['POL'] = df_temp['POL'].apply(standardize_pol)  # Standardize POL names
                all_pols.update(df_temp['POL'].dropna().astype(str).str.strip().unique())
    
    # Fetch custom vendors and process their most recent month-year document (without order_by)
    custom_vendors_ref = db.collection('custom_vendors')
    for doc in custom_vendors_ref.stream():
        vendor = doc.id
        if vendor not in data:  # Avoid overwriting hardcoded vendors
            data[vendor] = {}
            # Fetch all documents and manually find the most recent
            docs = vendors_ref.document(vendor).collection('data').stream()
            month_years = [(doc.id, doc.to_dict()) for doc in docs]
            if month_years:
                # Sort by month-year ID (e.g., "2025-MAY") in descending order
                most_recent = sorted(month_years, key=lambda x: x[0], reverse=True)[0]
                month_year, doc_data = most_recent
                data[vendor][month_year] = doc_data
                df_temp = pd.DataFrame(doc_data.get("data", []))
                if 'PORT' in df_temp.columns:
                    all_pods.update(df_temp['PORT'].dropna().astype(str).str.strip().unique())
                if 'POD' in df_temp.columns:
                    all_pods.update(df_temp['POD'].dropna().astype(str).str.strip().unique())
                if 'POL' in df_temp.columns:
                    df_temp['POL'] = df_temp['POL'].apply(standardize_pol)
                    all_pols.update(df_temp['POL'].dropna().astype(str).str.strip().unique())
    
    return data, sorted(list(all_pods)), sorted(list(all_pols))

def save_to_firestore(vendor, month_year, df, info):
    # Standardize POL names before saving
    if 'POL' in df.columns:
        df['POL'] = df['POL'].apply(standardize_pol)
    doc_ref = db.collection('vendors').document(vendor).collection('data').document(month_year)
    batch = db.batch()
    data_dict = {'data': df.to_dict(orient='records'), 'info': info}
    batch.set(doc_ref, data_dict, merge=True)
    batch.commit()

def save_custom_vendor_metadata(vendor_name, selected_columns):
    doc_ref = db.collection('custom_vendors').document(vendor_name)
    doc_ref.set({
        'name': vendor_name,
        'columns': selected_columns,
        'created_at': firestore.SERVER_TIMESTAMP
    })

def get_all_vendors():
    # Hardcoded vendors from load_from_firestore
    data, _, _ = load_from_firestore()
    vendors = list(data.keys())
    
    # Add custom vendors
    custom_vendors_ref = db.collection('custom_vendors')
    custom_docs = custom_vendors_ref.stream()
    for doc in custom_docs:
        vendor_name = doc.id
        if vendor_name not in vendors:
            vendors.append(vendor_name)
    
    return sorted(vendors)

def query_firestore(pol, pod, equipment, month=None, year=None):
    results = []
    vendors = get_all_vendors()
    
    # Get current month and year from system date
    current_date = datetime.now()
    current_year = str(current_date.year)  # e.g., "2025"
    current_month = current_date.strftime("%b").upper()  # e.g., "MAY"
    
    for vendor in vendors:
        vendor_ref = db.collection('vendors').document(vendor).collection('data')
        docs = vendor_ref.stream()
        for doc in docs:
            # doc.id is in the format "YYYY-MMM" (e.g., "2025-MAY")
            doc_month_year = doc.id.split('-')
            doc_year = doc_month_year[0]  # e.g., "2025"
            doc_month = doc_month_year[1]  # e.g., "MAY"
            
            # Use current month and year if not specified
            target_year = year if year else current_year
            target_month = month if month else current_month
            
            if doc_year != target_year or doc_month != target_month:
                continue
                
            doc_data = doc.to_dict()
            df = pd.DataFrame(doc_data.get("data", []))
            if df.empty:
                continue

            # Standardize POL names in the DataFrame
            if 'POL' in df.columns:
                df['POL'] = df['POL'].apply(standardize_pol)

            # Clean PORT/POD columns
            if 'PORT' in df.columns:
                df['PORT'] = df['PORT'].astype(str).str.strip()
                df['PORT'] = df['PORT'].replace(['nan', 'None', ''], pd.NA)
            if 'POD' in df.columns:
                df['POD'] = df['POD'].astype(str).str.strip()
                df['POD'] = df['POD'].replace(['nan', 'None', ''], pd.NA)

            # Apply POL filter with exact match
            if 'POL' in df.columns and pol:
                df = df[df['POL'].str.lower() == pol.lower()]

            # Apply POD filter with partial match
            if 'PORT' in df.columns and pod:
                df = df[df['PORT'].str.contains(pod, case=False, na=False)]
                df = df[df['PORT'].ne('') & df['PORT'].ne('None')]
            elif 'POD' in df.columns and pod:
                df = df[df['POD'].str.contains(pod, case=False, na=False)]
                df = df[df['POD'].ne('') & df['POD'].ne('None')]
                if not df.empty:
                    df = df.rename(columns={'POD': 'PORT'})

            # Apply equipment filter
            if equipment and any(eq in df.columns for eq in equipment):
                conditions = []
                for eq in equipment:
                    if eq in df.columns:
                        df[eq] = df[eq].astype(str).replace(['nan', 'None', ''], pd.NA)
                        condition = df[eq].notna() & (df[eq] != '') & (df[eq] != 'None')
                        conditions.append(condition)
                if conditions:
                    df = df[pd.concat([pd.Series(c) for c in conditions], axis=1).any(axis=1)]

            # Additional equipment check
            if equipment:
                for eq in equipment:
                    if eq in df.columns:
                        df = df[df[eq].ne('None')]

            # Append results if DataFrame has valid data
            if not df.empty:
                if 'PORT' in df.columns:
                    df = df[df['PORT'].ne('None')]
                if df.empty:
                    continue

                df['Month-Year'] = doc.id
                df['Vendor'] = vendor
                results.append(df)

    return pd.concat(results, ignore_index=True) if results else pd.DataFrame()

def delete_from_firestore(vendor, month_year):
    db.collection('vendors').document(vendor).collection('data').document(month_year).delete()

def get_all_month_years():
    """Fetch all available Month-Year values from Firestore."""
    month_years = set()
    vendors_ref = db.collection('vendors')
    for vendor_doc in vendors_ref.stream():
        docs = vendors_ref.document(vendor_doc.id).collection('data').stream()
        for doc in docs:
            month_years.add(doc.id)
    return sorted(list(month_years), reverse=True)

# === STREAMLIT APP ===
st.set_page_config(page_title="Vendor Data Parser", layout="wide")

# Apply custom CSS with sophisticated color theme
st.markdown(
    """
    <style>
    .stApp {background: linear-gradient(135deg, #f0f4f8, #ffffff); color: #2c3e50; font-family: 'Helvetica Neue', sans-serif;}
    .stTitle {text-align: center; font-size: 2.5em; color: #1a252f; text-shadow: 1px 1px 3px rgba(0,0,0,0.1); margin-bottom: 30px;}
    .stSubheader {color: #34495e; font-weight: 600; margin-bottom: 15px; text-transform: uppercase; letter-spacing: 1px;}
    .section-container {background: #ffffff; border-radius: 12px; padding: 20px; margin-bottom: 20px; box-shadow: 0 4px 8px rgba(0,0,0,0.1); border: 1px solid #ecf0f1;}
    .stButton>button {background: linear-gradient(45deg, #3498db, #8e44ad); color: white; border: none; padding: 10px 20px; border-radius: 6px; font-weight: 500; transition: all 0.3s ease; box-shadow: 0 2px 4px rgba(0,0,0,0.1);}
    .stButton>button:hover {background: linear-gradient(45deg, #2980b9, #8e44ad); transform: translateY(-2px); box-shadow: 0 4px 8px rgba(0,0,0,0.2);}
    .stSelectbox, .stFileUploader {background-color: #ffffff; border: 1px solid #bdc3c7; border-radius: 6px; padding: 8px; box-shadow: inset 0 1px 2px rgba(0,0,0,0.05);}
    .stSelectbox div[data-baseweb="select"] {border: 1px solid #bdc3c7; border-radius: 6px;}
    .stCheckbox {margin-right: 15px;}
    .equipment-container {display: flex; align-items: center; margin-top: 5px;}
    .equipment-container label {margin-right: 20px; font-size: 14px; color: #34495e;}
    .search-row {display: flex; align-items: center; gap: 15px;}
    .search-row .stSelectbox {flex: 1; min-width: 150px;}
    .search-row .equipment-container {flex: 1; min-width: 300px;}
    .search-row .stButton {flex: 0;}
    .vendor-button {width: 100%; text-align: center; background: #ecf0f1; border-radius: 6px; padding: 10px; transition: background 0.3s ease;}
    .vendor-button:hover {background: #dfe6ea;}
    .horizontal-checkboxes {display: flex; gap: 20px; align-items: center;}
    .horizontal-checkboxes label {margin-right: 0; font-size: 14px; color: #34495e;}
    </style>
    """,
    unsafe_allow_html=True
)

# Initialize session state
if 'page' not in st.session_state:
    st.session_state.page = 'main'
if 'selected_vendor' not in st.session_state:
    st.session_state.selected_vendor = None
if 'pod_suggestions' not in st.session_state or 'pol_suggestions' not in st.session_state:
    _, st.session_state.pod_suggestions, pol_suggestions_from_db = load_from_firestore()
    st.session_state.pod_suggestions = sorted(st.session_state.pod_suggestions)
    mandatory_pols = {"Nhava Sheva", "Rajula", "Pipavav"}
    st.session_state.pol_suggestions = sorted(list(set(pol_suggestions_from_db) | mandatory_pols))
if 'search_results' not in st.session_state:
    st.session_state.search_results = pd.DataFrame()

def delete_vendor_from_firestore(vendor):
    try:
        # Delete from vendors collection
        vendor_ref = db.collection('vendors').document(vendor)
        data_ref = vendor_ref.collection('data').stream()
        batch = db.batch()
        for doc in data_ref:
            batch.delete(doc.reference)
        batch.delete(vendor_ref)
        batch.commit()

        # Delete from custom_vendors collection (if exists)
        custom_vendor_ref = db.collection('custom_vendors').document(vendor)
        if custom_vendor_ref.get().exists:
            custom_vendor_ref.delete()

        # Clear cache and update suggestions
        st.cache_data.clear()
        data, pod_suggestions, pol_suggestions_from_db = load_from_firestore()
        st.session_state.pod_suggestions = sorted(pod_suggestions)
        mandatory_pols = {"Nhava Sheva", "Rajula", "Pipavav"}
        st.session_state.pol_suggestions = sorted(list(set(pol_suggestions_from_db) | mandatory_pols))
        st.success(f"Vendor '{vendor}' deleted successfully.")
    except Exception as e:
        st.error(f"Error deleting vendor '{vendor}': {str(e)}")

def main_page():
    st.title("Vendor Data Parser")

    # Sidebar Navigation with Clickable Buttons
    st.sidebar.title("Navigation")
    
    # Initialize selected_section in session state if not present
    if 'selected_section' not in st.session_state:
        st.session_state.selected_section = "Search Vendor Data"
    
    sections = ["Search Vendor Data","Upload Vendor Data", "Vendors", "Add New Vendor", "Remove Vendor"]
    
    # Create clickable buttons for each section
    for section in sections:
        if st.sidebar.button(section, key=f"nav_{section}", use_container_width=True):
            st.session_state.selected_section = section
            st.rerun()
    
    # Style the buttons to look like clickable text links
    st.sidebar.markdown("""
        <style>
        /* Style the sidebar buttons to look like clickable text */
        div[data-testid="stSidebar"] button {
            background: none;
            border: none;
            padding: 8px 0;
            color: #4A4A4A;
            font-size: 16px;
            text-align: left;
            width: 100%;
            cursor: pointer;
        }
        /* Highlight the selected section */
        div[data-testid="stSidebar"] button[kind="secondary"][aria-selected="true"] {
            color: #FF4B4B;
            font-weight: bold;
        }
        /* Add hover effect */
        div[data-testid="stSidebar"] button:hover {
            color: #FF4B4B;
            background: none;
        }
        </style>
    """, unsafe_allow_html=True)
    
    # Use the selected section from session state
    selected_section = st.session_state.selected_section

    if selected_section == "Upload Vendor Data":
        with st.container():
            st.subheader("Upload Vendor Data")
            col1, col2, col3 = st.columns(3)
            with col1:
                month = st.selectbox("Month", ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])
            with col2:
                year = st.selectbox("Year", [2025, 2026, 2027, 2028, 2029, 2030], format_func=lambda x: str(x))
            with col3:
                vendors = get_all_vendors()
                vendor = st.selectbox("Vendor", vendors)
            month_year = f"{year}-{month[:3].upper()}"

            if 'file_uploader_key' not in st.session_state:
                st.session_state.file_uploader_key = 0
            uploaded_file = st.file_uploader("Upload Excel or PDF", type=['xlsx', 'pdf'], key=f"uploader_{st.session_state.file_uploader_key}")

            if uploaded_file and st.button("Parse and Store"):
                parser_map = {
                    "MSC": parse_msc,
                    "Wan Hai": parse_wan_hai,
                    "Emirates": parse_emirates,
                    "ONE MRG": parse_one,
                    "HMM MRG": parse_hmm,
                    "OOCL": parse_oocl,
                    "PIL MRG": parse_pil,
                    "ARKAS MRG": parse_arkas,
                    "Interasia": parse_interasia,
                    "Cosco-Gulf": parse_cosco_gulf,
                    "Cosco-WCSA & CB": parse_cosco_wcsa_cb,
                    "Cosco-Africa": parse_cosco_africa,
                    "Turkon": parse_turkon,
                    "Cosco-Fareast": parse_cosco_fareast,
                    "ZIM MRG": parse_zim,
                    "MSC-EUR MED": parse_msc_eur_med
                }
                try:
                    custom_vendors_ref = db.collection('custom_vendors')
                    custom_doc = custom_vendors_ref.document(vendor).get()
                    if custom_doc.exists:
                        selected_columns = custom_doc.to_dict().get('columns', ['POL', 'PORT', '20', '40STD', '40HC', 'REMARKS'])
                        df, info = parse_custom_vendor(uploaded_file, month_year, selected_columns)
                    else:
                        df, info = parser_map[vendor](uploaded_file, month_year)
                    if not df.empty:
                        save_to_firestore(vendor, month_year, df, info)
                        if custom_doc.exists:
                            save_custom_vendor_metadata(vendor, selected_columns)
                        st.cache_data.clear()
                        data, pod_suggestions, pol_suggestions_from_db = load_from_firestore()
                        st.session_state.pod_suggestions = sorted(pod_suggestions)
                        mandatory_pols = {"Nhava Sheva", "Rajula", "Pipavav"}
                        st.session_state.pol_suggestions = sorted(list(set(pol_suggestions_from_db) | mandatory_pols))
                        st.success(f"Data for {vendor} ({month_year}) parsed and stored successfully! Overwritten if previously existed.")
                        st.session_state.file_uploader_key += 1
                        st.rerun()
                    else:
                        st.warning(f"No data extracted for {vendor} ({month_year}). Check file format.")
                except Exception as e:
                    st.error(f"Error parsing file: {str(e)}")
                    st.session_state.file_uploader_key += 1
                    st.rerun()

    elif selected_section == "Vendors":
        with st.container():
            st.subheader("Vendors")
            data, _, _ = load_from_firestore()
            vendors = get_all_vendors()
            num_columns = 5
            num_rows = (len(vendors) + num_columns - 1) // num_columns
            for row_idx in range(num_rows):
                cols = st.columns(num_columns)
                for col_idx in range(num_columns):
                    vendor_idx = row_idx * num_columns + col_idx
                    if vendor_idx < len(vendors):
                        with cols[col_idx]:
                            vendor = vendors[vendor_idx]
                            record_count = len(data.get(vendor, {}))
                            button_text = f"{vendor} ({record_count} records)"
                            if st.button(button_text, key=f"vendor_{vendor}", help=f"View {vendor} data", use_container_width=True):
                                st.session_state.page = 'vendor'
                                st.session_state.selected_vendor = vendor
                                st.rerun()

    elif selected_section == "Add New Vendor":
        with st.container():
            st.subheader("Add New Vendor")
            col1, col2 = st.columns([3, 1])
            with col1:
                new_vendor_name = st.text_input("Vendor Name", help="Enter a unique name for the new vendor")
            with col2:
                st.write("")
                add_vendor_button = st.button("Add Vendor")
            
            possible_columns = ['POL', 'PORT', '20', '40STD', '40HC', 'REMARKS', 'ROUTING', 'TRANSIT TIME', 'VALIDITY', 'SERVICE', 'EXPIRY DATE']
            selected_columns = st.multiselect("Select Column Headings", possible_columns, default=['POL', 'PORT'], help="Select the columns to include for the new vendor.")

            if add_vendor_button and new_vendor_name:
                vendors = get_all_vendors()
                if new_vendor_name in vendors:
                    st.error(f"Vendor '{new_vendor_name}' already exists. Choose a different name.")
                elif not new_vendor_name.strip():
                    st.error("Vendor name cannot be empty.")
                elif not selected_columns:
                    st.error("Please select at least one column for the new vendor.")
                else:
                    with st.spinner(f"Adding vendor '{new_vendor_name}'..."):
                        save_custom_vendor_metadata(new_vendor_name, selected_columns)
                    st.success(f"Vendor '{new_vendor_name}' added successfully!")
                    st.rerun()

            st.write("Download an Excel template with the selected column headings.")
            excel_buffer = io.BytesIO()
            if new_vendor_name.strip() and selected_columns:
                df = pd.DataFrame(columns=selected_columns)
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=new_vendor_name, index=False)
                excel_buffer.seek(0)
            else:
                excel_buffer = None

            st.download_button(
                label="Download Template",
                data=excel_buffer if excel_buffer else io.BytesIO(),
                file_name=f"{new_vendor_name}_template.xlsx" if new_vendor_name.strip() else "template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                disabled=not (new_vendor_name.strip() and selected_columns),
                help="Click to download an Excel template with the selected column headings."
            )

    elif selected_section == "Remove Vendor":
        with st.container():
            st.subheader("Remove Vendor")
            col1, col2 = st.columns([3, 1])
            with col1:
                vendors = get_all_vendors()
                vendor_to_remove = st.selectbox("Select Vendor to Remove", [""] + vendors, format_func=lambda x: "" if x == "" else x, help="Select a vendor to remove")
            with col2:
                st.write("")
                remove_vendor_button = st.button("Remove Vendor")
            
            if remove_vendor_button and vendor_to_remove:
                if not vendor_to_remove:
                    st.error("Please select a vendor to remove.")
                else:
                    with st.spinner(f"Removing vendor '{vendor_to_remove}'..."):
                        delete_vendor_from_firestore(vendor_to_remove)
                    st.rerun()

    elif selected_section == "Search Vendor Data":
        with st.container():
            st.subheader("Search Vendor Data")
            # Adjust column widths to fit POL, POD/PORT, Carrier, Month, Year, Equipment, and Search button in one row
            col1, col2, col3, col4, col5, col6, col7 = st.columns([1.5, 1.5, 1.5, 1, 1, 1, 0.5])

            current_date = datetime.now()
            current_month = current_date.strftime("%b")
            current_year = str(current_date.year)

            #Initialize session state for POD input
            if 'pod_input_text' not in st.session_state:
                st.session_state.pod_input_text = ""
            if 'filtered_pod_suggestions' not in st.session_state:
                st.session_state.filtered_pod_suggestions = st.session_state.pod_suggestions
            if 'pod_selected' not in st.session_state:
                st.session_state.pod_selected = ""
            with col1:
                pol_input = st.selectbox("POL", [""] + st.session_state.pol_suggestions,
                                        format_func=lambda x: "" if x == "" else x, help="Start typing to see suggestions")
            with col2:
                all_ports = sorted(set(st.session_state.get("pod_suggestions", [])))

                # ✅ Initialize filtered suggestions if not already present
                if "filtered_pod_suggestions" not in st.session_state:
                    st.session_state.filtered_pod_suggestions = all_ports

                # ✅ Display dropdown only for filtered matches
                pod_input = st.selectbox(
                    "POD/PORT",
                    [""] + st.session_state.filtered_pod_suggestions,
                    format_func=lambda x: "" if x == "" else x,
                    key="pod_selectbox",
                    help="Select a POD/PORT (exact match only)",
                    index=0
                )

                # ✅ On selection change, update filtered suggestions using exact match logic
                if pod_input != st.session_state.get("pod_selected", ""):
                    st.session_state.pod_selected = pod_input

                    if pod_input.strip():
                        input_lower = pod_input.strip().lower()

                        # ✅ Exact match filtering (case-insensitive)
                        filtered_suggestions = [
                            pod for pod in st.session_state.pod_suggestions
                            if pod and pod.strip() and pod.lower() == input_lower
                        ]

                        st.session_state.filtered_pod_suggestions = sorted(filtered_suggestions)
                    else:
                        st.session_state.filtered_pod_suggestions = all_ports

                    st.rerun()


            with col3:
                data, _, _ = load_from_firestore()
                vendors = sorted(data.keys())  # Extract vendor names from the data dictionary
                carrier_input = st.selectbox("Carrier", [""] + vendors,
                                            format_func=lambda x: "" if x == "" else x,
                                            help="Start typing to see vendor names")
            with col4:
                months = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
                month_index = months.index(current_month) if current_month in months else 0
                month_input = st.selectbox("Month", months, index=month_index,
                                           format_func=lambda x: "" if x == "" else x, help="Select a month")
            with col5:
                years = [""] + [str(y) for y in range(2025, 2031)]
                year_index = years.index(current_year) if current_year in years else 0
                year_input = st.selectbox("Year", years, index=year_index,
                                          format_func=lambda x: "" if x == "" else x, help="Select a year")
            with col6:
                equipment_options = ["20", "40STD", "40HC"]
                equipment = st.multiselect("Equipment", equipment_options, default=[],
                                           help="Select one or more equipment types")
            with col7:
                st.markdown("<br>", unsafe_allow_html=True)  # Add vertical alignment
                if st.button("🔍", help="Search", use_container_width=True):
                    if not (pol_input or carrier_input):
                        st.warning("Please select at least one of POL or Carrier.")
                    elif not (pod_input or equipment):
                        st.warning("Please enter a POD/PORT or select at least one equipment type.")
                    else:
                        month = month_input[:3].upper() if month_input else None
                        year = year_input if year_input else None
                        results = query_firestore(pol_input, pod_input, equipment, month=month, year=year)
                        if carrier_input:
                            results = results[results['Vendor'].str.lower() == carrier_input.lower()] if not results.empty else results
                        st.session_state.search_results = results
                        if st.session_state.search_results.empty:
                            st.info("No matching records found.")

            if not st.session_state.search_results.empty:
                st.subheader("Search Results")
                display_columns = ['Vendor', 'Month-Year', 'POL', 'PORT'] + \
                                [e for e in equipment if e in st.session_state.search_results.columns]
                additional_cols = ['REMARKS', 'ROUTING', 'TRANSIT TIME', 'VALIDITY', 'SERVICE NAME',
                                'SERVICE', 'EXPIRY DATE', 'Sheet', 'IMO SC PER TEU', '20 Haz', '40 Haz', "40'HRF"]
                display_columns.extend([c for c in additional_cols if c in st.session_state.search_results.columns])

                if 'POL' in st.session_state.search_results.columns:
                    st.session_state.search_results['POL'] = st.session_state.search_results['POL'].apply(standardize_pol)
                if 'POD' in st.session_state.search_results.columns:
                    st.session_state.search_results = st.session_state.search_results.rename(columns={'POD': 'PORT'})

                st.dataframe(st.session_state.search_results[display_columns])

def vendor_page():
    vendor = st.session_state.selected_vendor
    st.title(f"{vendor} Data")

    if st.button("Back to Main"):
        st.session_state.page = 'main'
        st.session_state.selected_vendor = None
        st.rerun()

    data, _, _ = load_from_firestore()
    if vendor not in data or not data[vendor]:
        st.info(f"No data available for {vendor}.")
        return

    def standardize_currency(value):
        if pd.isna(value):
            return value
        value = str(value).strip()
        # Match formats like "100$", "$100", "100 €", "€200", etc.
        match = re.match(r'^\s*([\$€])?\s*(\d*\.?\d+)\s*([\$€])?\s*$', value)
        if match:
            currency_before, number, currency_after = match.groups()
            if currency_before:
                return f"{currency_before}{number}"
            elif currency_after:
                return f"{currency_after}{number}"
        return value

    for month_year in sorted(data[vendor].keys(), reverse=True):
        with st.expander(f"{month_year[:4]} {month_year[5:]}"):
            df = pd.DataFrame(data[vendor][month_year]["data"])
            info = data[vendor][month_year]["info"]

            if not df.empty:
                text_columns = ['REMARKS', 'ROUTING', 'TRANSIT TIME', 'VALIDITY', 'SERVICE NAME', 'SERVICE', 'EXPIRY DATE']
                df_values = set()
                for col in text_columns:
                    if col in df.columns:
                        df_values.update(df[col].dropna().apply(normalize_text))
                filtered_info = [line for line in info if normalize_text(line) not in df_values and line.strip()]

                # Display Data section first
                st.subheader("Data")
                # Ensure POL is displayed consistently and rename POD to PORT for HMM MRG
                if 'POL' in df.columns:
                    df['POL'] = df['POL'].apply(standardize_pol)
                if vendor == "HMM MRG" and 'POD' in df.columns:
                    df = df.rename(columns={'POD': 'PORT'})
                
                # Standardize currency formats in cost-related columns
                cost_columns = ['20', '20 Haz', '40STD', '40HC', '40 Haz', 'IMO SC PER TEU', "40'HRF"]
                for col in cost_columns:
                    if col in df.columns:
                        df[col] = df[col].apply(standardize_currency)
                
                # Standardize column names
                column_mapping = {
                    'Routing': 'ROUTING',
                    'SERVICE NAME': 'SERVICE',
                    'Description': 'REMARKS',
                    'Remarks': 'REMARKS'
                }
                df = df.rename(columns=column_mapping)
                
                # Define the desired column order
                desired_columns = ['POL', 'PORT', '20', '40STD', '40HC', 'ROUTING', 'TRANSIT TIME', 'SERVICE', 'EXPIRY DATE', 'VALIDITY', 'REMARKS']
                
                # Handle related columns (20 Haz after 20, 40 Haz after 40HC)
                final_columns = []
                for col in desired_columns:
                    final_columns.append(col)
                    if col == '20' and '20 Haz' in df.columns:
                        final_columns.append('20 Haz')
                    if col == '40HC' and '40 Haz' in df.columns:
                        final_columns.append('40 Haz')
                
                # Add any additional columns (not in desired_columns or related columns) before REMARKS
                additional_columns = [col for col in df.columns if col not in final_columns]
                if additional_columns:
                    remarks_idx = final_columns.index('REMARKS')
                    final_columns = final_columns[:remarks_idx] + additional_columns + final_columns[remarks_idx:]
                
                # Filter columns that exist in the DataFrame
                final_columns = [col for col in final_columns if col in df.columns]
                
                # Reorder DataFrame columns
                df = df[final_columns]

                st.dataframe(df)

                # Excel download button (after Data section, before Additional Information)
                output_file = f"{vendor}_{month_year}.xlsx"
                excel_buffer = io.BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    start_row = 0
                    if filtered_info:
                        ws = writer.book.create_sheet(vendor)
                        for i, line in enumerate(filtered_info, start=1):
                            ws.cell(row=i, column=1, value=line)
                        start_row = len(filtered_info) + 1
                    df.to_excel(writer, sheet_name=vendor, startrow=start_row, index=False)
                    if vendor == "HMM MRG":
                        ws = writer.book[vendor]
                        for row in ws.iter_rows(min_row=start_row + 2, min_col=2, max_col=3):
                            for cell in row:
                                if cell.value and isinstance(cell.value, (int, float)):
                                    cell.value = f"${cell.value}"
                                cell.number_format = '@'  # Treat as text to preserve $ prefix
                    if vendor == "ONE MRG" and 'EXPIRY DATE' in df.columns:
                        ws = writer.book[vendor]
                        if start_row == 0:
                            start_row = 6
                            ws.insert_rows(1, amount=6)
                        for i, line in enumerate([l for l in filtered_info if '\n' in l][:6], start=1):
                            ws.cell(row=i, column=1, value=sanitize_text(line))

                excel_buffer.seek(0)

                st.download_button(
                    label=f"Download {month_year} Data",
                    data=excel_buffer,
                    file_name=output_file,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                # Display Additional Information section after the table
                if filtered_info or ('Sheet' in df.columns and df['Sheet'].dropna().unique().size > 0):
                    st.subheader("Additional Information")
                    if 'Sheet' in df.columns:
                        unique_sheets = df['Sheet'].dropna().unique()
                        if len(unique_sheets) > 0:
                            st.write("**Sheets:**")
                            for sheet in unique_sheets:
                                st.write(f"- {sheet}")
                    if filtered_info:
                        st.write("**Details:**")
                        for line in filtered_info:
                            st.write(f"- {line}")
            else:
                filtered_info = info
                st.warning("No data available for this period.")
                # Still display Additional Information if it exists
                if filtered_info or ('Sheet' in df.columns and df['Sheet'].dropna().unique().size > 0):
                    st.subheader("Additional Information")
                    if 'Sheet' in df.columns:
                        unique_sheets = df['Sheet'].dropna().unique()
                        if len(unique_sheets) > 0:
                            st.write("**Sheets:**")
                            for sheet in unique_sheets:
                                st.write(f"- {sheet}")
                    if filtered_info:
                        st.write("**Details:**")
                        for line in filtered_info:
                            st.write(f"- {line}")

            if st.button(f"Delete {month_year}", key=f"delete_{month_year}"):
                delete_from_firestore(vendor, month_year)
                st.cache_data.clear()
                st.success(f"Deleted {month_year} data for {vendor}.")
                st.rerun()

def add_vendor_page():
    st.title("Add New Vendor")
    vendor_name = st.session_state.new_vendor_name

    # Column selection
    available_columns = [
        'POL', 'PORT', '20', '40STD', '40HC', 'REMARKS', 'ROUTING',
        'TRANSIT TIME', 'SERVICE', 'VALIDITY', 'EXPIRY DATE'
    ]
    selected_columns = st.multiselect(
        "Select Columns (POL and PORT are required)",
        available_columns,
        default=['POL', 'PORT'],
        help="Select the columns present in your Excel file. POL and PORT are mandatory."
    )

    # Month and Year selection
    col1, col2 = st.columns(2)
    with col1:
        month = st.selectbox("Month", ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])
    with col2:
        year = st.selectbox("Year", [2025, 2026, 2027, 2028, 2029, 2030], format_func=lambda x: str(x))
    month_year = f"{year}-{month[:3].upper()}"

    # File uploader
    uploaded_file = st.file_uploader("Upload Excel File", type=['xlsx'], key=f"new_vendor_uploader")

    # Buttons
    col3, col4 = st.columns([1, 1])
    with col3:
        if st.button("Save Vendor and Parse Data"):
            if not uploaded_file:
                st.error("Please upload an Excel file.")
            elif 'POL' not in selected_columns or 'PORT' not in selected_columns:
                st.error("POL and PORT are required columns.")
            else:
                try:
                    # Parse the file
                    df, info = parse_custom_vendor(uploaded_file, month_year, selected_columns)
                    if not df.empty:
                        # Save vendor metadata and data
                        save_custom_vendor_metadata(vendor_name, selected_columns)
                        save_to_firestore(vendor_name, month_year, df, info)
                        # Update suggestions
                        st.cache_data.clear()
                        data, pod_suggestions, pol_suggestions_from_db = load_from_firestore()
                        st.session_state.pod_suggestions = sorted(pod_suggestions)
                        mandatory_pols = {"Nhava Sheva", "Rajula", "Pipavav"}
                        st.session_state.pol_suggestions = sorted(list(set(pol_suggestions_from_db) | mandatory_pols))
                        st.success(f"Vendor '{vendor_name}' added and data for {month_year} stored successfully!")
                        st.session_state.page = 'main'
                        st.session_state.new_vendor_name = None
                        st.rerun()
                    else:
                        st.warning("No data extracted from the file. Check the file format and column mappings.")
                except Exception as e:
                    st.error(f"Error processing file: {str(e)}")
    with col4:
        if st.button("Cancel"):
            st.session_state.page = 'main'
            st.session_state.new_vendor_name = None
            st.rerun()
if st.session_state.page == 'main':
    main_page()
elif st.session_state.page == 'add_vendor':
    add_vendor_page()
else:
    vendor_page()